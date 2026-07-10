"""
LC Waikiki - Multi-Depot Capacitated Vehicle Routing Problem (MDCVRP)
Google OR-Tools ile Sezgisel Optimizasyon Motoru
"""

import sys
import time
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp

# ─────────────────────────────────────────────
# 1. VERİ YÜKLEME
# ─────────────────────────────────────────────

EXCEL_PATH      = r"C:\Users\Zahid Sami\gams_output\01_Gercek_Veriler\LCW_Mesafe_Sure_Matrisi.xlsx"
SHEET_NAME      = "Mesafe (km)"
SHEET_NAME_SURE = "Süre (dk)"

NUM_DEPOTS  = 6       # İlk 6 indeks (0-5) depo
NUM_VEHICLES = 10
VEHICLE_CAPACITY = 3000
STORE_DEMAND = 150
DEPOT_DEMAND = 0
TIME_LIMIT_SECONDS = 60

# ─── MALİYET PARAMETRELERİ ──────────────────────────────────────────────────
YAKIT_FIYATI_TL         = 43.0
YAKIT_TUKETIM_LT_KM     = 0.35
ARAC_GUNLUK_SABIT_TL    = 2500
SURUCU_GUNLUK_TL        = 1500
OUTSOURCE_KOLI_FIYAT_TL = 3.75
TOPLAM_MAGAZA_SAYISI    = 533
INHOUSE_MAGAZA_SAYISI   = 198
OUTSOURCE_MAGAZA_SAYISI = 335

# Her araç hangi depodan çıkıp hangi depoya dönüyor
VEHICLE_STARTS = [0, 0, 1, 1, 2, 2, 3, 3, 4, 5]
VEHICLE_ENDS   = [0, 0, 1, 1, 2, 2, 3, 3, 4, 5]


def yukle_mesafe_matrisi(path: str, sheet: str) -> list[list[int]]:
    """Excel dosyasından mesafe matrisini okur ve metre cinsinden int'e çevirir."""
    print(f"  Dosya okunuyor: {path}")
    df = pd.read_excel(path, sheet_name=sheet, index_col=0)
    print(f"  Matris boyutu: {df.shape[0]} x {df.shape[1]}")

    if df.shape != (204, 204):
        print(f"  [UYARI] Beklenen boyut 204x204, okunan: {df.shape}")

    # km → metre, float → int
    matrix = (df.fillna(0) * 1000).astype(int).values.tolist()
    return matrix


def yukle_sure_matrisi(path: str, sheet: str) -> list[list[int]]:
    """Excel dosyasından süre matrisini okur; dakika cinsinden int döndürür."""
    print(f"  Süre matrisi okunuyor: '{sheet}' sayfası")
    df = pd.read_excel(path, sheet_name=sheet, index_col=0)
    print(f"  Süre matrisi boyutu: {df.shape[0]} x {df.shape[1]}")

    if df.shape != (204, 204):
        print(f"  [UYARI] Beklenen boyut 204x204, okunan: {df.shape}")

    # Dakika, float → int (yuvarla)
    matrix = df.fillna(0).round().astype(int).values.tolist()
    return matrix


# ─────────────────────────────────────────────
# 2. OR-TOOLS VERİ MODELİ
# ─────────────────────────────────────────────

def veri_modeli_olustur(mesafe_matrisi: list[list[int]], sure_matrisi: list[list[int]]) -> dict:
    num_locations = len(mesafe_matrisi)
    num_stores    = num_locations - NUM_DEPOTS

    # Talep vektörü: depolar 0, mağazalar 150
    demands = [DEPOT_DEMAND] * NUM_DEPOTS + [STORE_DEMAND] * num_stores

    return {
        "distance_matrix":    mesafe_matrisi,
        "duration_matrix":    sure_matrisi,
        "demands":            demands,
        "vehicle_capacities": [VEHICLE_CAPACITY] * NUM_VEHICLES,
        "num_vehicles":       NUM_VEHICLES,
        "starts":             VEHICLE_STARTS,
        "ends":               VEHICLE_ENDS,
        "num_locations":      num_locations,
        "service_time":       15,   # dakika — mağaza başı hizmet süresi
    }


# ─────────────────────────────────────────────
# 3. ÇÖZÜCܒ
# ─────────────────────────────────────────────

def coz(data: dict):
    manager = pywrapcp.RoutingIndexManager(
        data["num_locations"],
        data["num_vehicles"],
        data["starts"],
        data["ends"],
    )
    routing = pywrapcp.RoutingModel(manager)

    # ── Mesafe callback ──
    def distance_callback(from_idx, to_idx):
        i = manager.IndexToNode(from_idx)
        j = manager.IndexToNode(to_idx)
        return data["distance_matrix"][i][j]

    transit_cb_idx = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_cb_idx)

    # ── Kapasite kısıtı ──
    def demand_callback(idx):
        node = manager.IndexToNode(idx)
        return data["demands"][node]

    demand_cb_idx = routing.RegisterUnaryTransitCallback(demand_callback)
    routing.AddDimensionWithVehicleCapacity(
        demand_cb_idx,
        0,                                  # slack yok
        data["vehicle_capacities"],
        True,                               # sıfırdan başla
        "Capacity",
    )

    # ── Arama parametreleri: Guided Local Search ──
    search_params = pywrapcp.DefaultRoutingSearchParameters()
    search_params.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.SAVINGS
    )
    search_params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    )
    search_params.time_limit.seconds = TIME_LIMIT_SECONDS
    search_params.log_search = False

    print(f"\n  Optimizasyon başlatılıyor ({TIME_LIMIT_SECONDS} sn zaman limiti)...")
    t0 = time.time()
    solution = routing.SolveWithParameters(search_params)
    elapsed = time.time() - t0
    print(f"  Tamamlandı: {elapsed:.1f} saniye")

    return manager, routing, solution


# ─────────────────────────────────────────────
# 4. RAPOR
# ─────────────────────────────────────────────

RENK = {
    "BASLIK":   "\033[1;36m",   # bold cyan
    "ARAC":     "\033[1;33m",   # bold yellow
    "DEPO":     "\033[1;32m",   # bold green
    "BILGI":    "\033[0;37m",   # white
    "TOPLAM":   "\033[1;35m",   # bold magenta
    "RESET":    "\033[0m",
    "CIZGI":    "\033[0;34m",   # blue
}

def _r(key: str, text: str) -> str:
    return f"{RENK[key]}{text}{RENK['RESET']}"


def rapor_yazdir(manager, routing, solution, data):
    if not solution:
        print("\n  [HATA] Çözüm bulunamadı!")
        return

    print()
    print(_r("CIZGI", "═" * 70))
    print(_r("BASLIK", "           LC WİKİKİ — ÖZET FİLO RAPORU"))
    print(_r("CIZGI", "═" * 70))

    capacity_dim = routing.GetDimensionOrDie("Capacity")
    toplam_mesafe = 0
    toplam_yuklu_arac = 0

    for v in range(data["num_vehicles"]):
        idx = routing.Start(v)
        # Araç çıkış deposu
        baslangic_depo = manager.IndexToNode(idx)

        # Rota düğümlerini topla
        rota_dugumler = []
        while not routing.IsEnd(idx):
            node = manager.IndexToNode(idx)
            rota_dugumler.append(node)
            idx = solution.Value(routing.NextVar(idx)
        )
        bitis_depo = manager.IndexToNode(idx)

        # Sadece depo → depo gidiş-dönüşü (mağaza yok) ise atla
        magaza_sayisi = sum(1 for n in rota_dugumler if n >= NUM_DEPOTS)
        if magaza_sayisi == 0:
            continue

        toplam_yuklu_arac += 1

        # Mesafe hesapla (metre → km)
        mesafe_m = solution.ObjectiveValue()   # genel toplam; araç bazlı hesap:
        mesafe_m_arac = 0
        idx2 = routing.Start(v)
        while not routing.IsEnd(idx2):
            next_idx = solution.Value(routing.NextVar(idx2))
            i = manager.IndexToNode(idx2)
            j = manager.IndexToNode(next_idx)
            mesafe_m_arac += data["distance_matrix"][i][j]
            idx2 = next_idx

        mesafe_km = mesafe_m_arac / 1000
        toplam_mesafe += mesafe_km

        # Yük
        son_idx = routing.End(v)
        yuk = solution.Value(capacity_dim.CumulVar(son_idx))

        # Mağaza listesi (depo düğümleri hariç)
        magaza_listesi = [str(n) for n in rota_dugumler if n >= NUM_DEPOTS]

        print()
        print(_r("ARAC", f"  🚛 ARAÇ {v + 1:>2d}"))
        print(_r("CIZGI", "  " + "─" * 66))
        print(_r("BILGI",  f"  {'Çıkış Deposu':<20}: ") + _r("DEPO",  f"Depo {baslangic_depo}"))
        print(_r("BILGI",  f"  {'Dönüş Deposu':<20}: ") + _r("DEPO",  f"Depo {bitis_depo}"))
        print(_r("BILGI",  f"  {'Uğranan Mağaza Sayısı':<20}: {magaza_sayisi}"))
        print(_r("BILGI",  f"  {'Toplam Yük (koli)':<20}: {yuk} / {VEHICLE_CAPACITY}"))
        print(_r("BILGI",  f"  {'Toplam Mesafe':<20}: {mesafe_km:.1f} km"))

        # Mağazaları 10'arlık satırlarda göster
        satirlar = [magaza_listesi[i:i+10] for i in range(0, len(magaza_listesi), 10)]
        for k, satir in enumerate(satirlar):
            etiket = "Mağaza Rotası" if k == 0 else ""
            print(_r("BILGI", f"  {etiket:<20}: ") + " → ".join(satir))

    print()
    print(_r("CIZGI", "═" * 70))
    print(_r("TOPLAM", f"  TOPLAM AKTİF ARAÇ  : {toplam_yuklu_arac} / {NUM_VEHICLES}"))
    print(_r("TOPLAM", f"  TOPLAM FILO MESAFESİ: {toplam_mesafe:.1f} km"))
    print(_r("TOPLAM", f"  ÇÖZÜM OBJEKTİF DEĞ : {solution.ObjectiveValue() / 1000:.1f} km"))
    print(_r("CIZGI", "═" * 70))
    print()

    # ── Maliyet Analizi (konsol özeti) ──
    yakit          = toplam_mesafe * YAKIT_TUKETIM_LT_KM * YAKIT_FIYATI_TL
    sabit          = toplam_yuklu_arac * ARAC_GUNLUK_SABIT_TL
    personel       = toplam_yuklu_arac * SURUCU_GUNLUK_TL
    inhouse_toplam = yakit + sabit + personel
    tam_outsource  = TOPLAM_MAGAZA_SAYISI * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    hibrit_out     = OUTSOURCE_MAGAZA_SAYISI * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    hibrit_toplam  = inhouse_toplam + hibrit_out
    tasarruf_gun   = tam_outsource - hibrit_toplam
    tasarruf_yil   = tasarruf_gun * 300
    tasarruf_oran  = (tasarruf_gun / tam_outsource) * 100

    print(_r("CIZGI",  "═" * 70))
    print(_r("BASLIK", "           LC WİKİKİ — MALİYET ANALİZİ"))
    print(_r("CIZGI",  "═" * 70))
    print(_r("BILGI",  f"  Yakıt Maliyeti           : {yakit:>12,.0f} TL/gün"))
    print(_r("BILGI",  f"  Araç Sabit Maliyet       : {sabit:>12,.0f} TL/gün"))
    print(_r("BILGI",  f"  Sürücü Personel          : {personel:>12,.0f} TL/gün"))
    print(_r("TOPLAM", f"  In-house Günlük Toplam   : {inhouse_toplam:>12,.0f} TL/gün"))
    print(_r("CIZGI",  "─" * 70))
    print(_r("BILGI",  f"  Tam Outsource (3PL)      : {tam_outsource:>12,.0f} TL/gün"))
    print(_r("BILGI",  f"  Hibrit Outsource Kısmı   : {hibrit_out:>12,.0f} TL/gün"))
    print(_r("TOPLAM", f"  Hibrit Günlük Toplam     : {hibrit_toplam:>12,.0f} TL/gün"))
    print(_r("CIZGI",  "─" * 70))
    print(_r("TOPLAM", f"  Günlük Tasarruf          : {tasarruf_gun:>12,.0f} TL/gün"))
    print(_r("TOPLAM", f"  Yıllık Tasarruf (300 gün): {tasarruf_yil:>12,.0f} TL/yıl"))
    print(_r("TOPLAM", f"  Tasarruf Oranı           : %{tasarruf_oran:>11.1f}"))
    print(_r("CIZGI",  "═" * 70))
    print()


# ─────────────────────────────────────────────
# 5. EXCEL EXPORT
# ─────────────────────────────────────────────

# Renk paleti (hex)
C_BASLIK_BG   = "1F3864"   # koyu lacivert
C_BASLIK_FG   = "FFFFFF"
C_DEPO_BG     = "2E75B6"   # mavi
C_DEPO_FG     = "FFFFFF"
C_ARAC_BG     = "D6E4F0"   # açık mavi
C_TOPLAM_BG   = "FFF2CC"   # sarı
C_TOPLAM_FG   = "7B3F00"
C_SATIR1      = "EBF3FB"
C_SATIR2      = "FFFFFF"
C_MAĞAZA_BG   = "E2EFDA"   # açık yeşil
C_KENARLIK    = "BDD7EE"


def _doldur(hex_kod: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_kod)


def _kenar() -> Border:
    ince = Side(style="thin", color=C_KENARLIK)
    return Border(left=ince, right=ince, top=ince, bottom=ince)


def _boyutlandir(ws, col_letter: str, genislik: float):
    ws.column_dimensions[col_letter].width = genislik


def excel_export(manager, routing, solution, data, cikti_yolu: str):
    if not solution:
        print("  [UYARI] Çözüm yok, Excel oluşturulmadı.")
        return

    wb = Workbook()

    # ── Sayfa 1: Özet Filo Raporu ──────────────────────────────────────
    ws_ozet = wb.active
    ws_ozet.title = "Özet Filo Raporu"

    baslik_font  = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=11)
    depo_font    = Font(name="Calibri", bold=True, color=C_DEPO_FG,   size=10)
    normal_font  = Font(name="Calibri", size=10)
    toplam_font  = Font(name="Calibri", bold=True, color=C_TOPLAM_FG, size=11)

    orta = Alignment(horizontal="center", vertical="center")
    sol  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Başlık satırı
    ws_ozet.merge_cells("A1:J1")
    c = ws_ozet["A1"]
    c.value         = "LC WİKİKİ — MDCVRP ÖZET FİLO RAPORU"
    c.font          = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=14)
    c.fill          = _doldur(C_BASLIK_BG)
    c.alignment     = orta
    ws_ozet.row_dimensions[1].height = 32

    # Tarih satırı
    ws_ozet.merge_cells("A2:J2")
    c2 = ws_ozet["A2"]
    c2.value     = f"Oluşturma Tarihi: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c2.font      = Font(name="Calibri", italic=True, color="666666", size=9)
    c2.fill      = _doldur("D6E4F0")
    c2.alignment = orta
    ws_ozet.row_dimensions[2].height = 16

    # Sütun başlıkları
    basliklar = ["Araç", "Çıkış Deposu", "Dönüş Deposu",
                 "Mağaza Sayısı", "Toplam Yük (koli)",
                 "Kapasite Doluluk %", "Toplam Mesafe (km)", "Mağaza Rotası",
                 "Toplam Süre (dk)", "540 dk Aşımı"]
    ws_ozet.row_dimensions[3].height = 20
    for col, bas in enumerate(basliklar, start=1):
        c = ws_ozet.cell(row=3, column=col, value=bas)
        c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill      = _doldur(C_DEPO_BG)
        c.alignment = orta
        c.border    = _kenar()

    capacity_dim = routing.GetDimensionOrDie("Capacity")
    satir = 4
    toplam_mesafe  = 0.0
    toplam_arac    = 0
    ozet_satirlar  = []  # Sayfa 2 için

    for v in range(data["num_vehicles"]):
        idx             = routing.Start(v)
        baslangic_depo  = manager.IndexToNode(idx)
        rota_dugumler   = []

        while not routing.IsEnd(idx):
            node = manager.IndexToNode(idx)
            rota_dugumler.append(node)
            idx = solution.Value(routing.NextVar(idx))

        bitis_depo    = manager.IndexToNode(idx)
        magaza_listesi = [n for n in rota_dugumler if n >= NUM_DEPOTS]

        if not magaza_listesi:
            continue

        toplam_arac += 1

        # Araç mesafesi
        mesafe_m = 0
        idx2 = routing.Start(v)
        while not routing.IsEnd(idx2):
            next_idx = solution.Value(routing.NextVar(idx2))
            i = manager.IndexToNode(idx2)
            j = manager.IndexToNode(next_idx)
            mesafe_m += data["distance_matrix"][i][j]
            idx2 = next_idx

        mesafe_km  = mesafe_m / 1000
        toplam_mesafe += mesafe_km
        yuk        = solution.Value(capacity_dim.CumulVar(routing.End(v)))
        doluluk    = round(yuk / VEHICLE_CAPACITY * 100, 1)
        rota_str   = " → ".join(str(n) for n in magaza_listesi)

        # Araç süresi: seyahat (dk) + mağaza başı 15 dk hizmet süresi
        sure_dk = 0
        idx3 = routing.Start(v)
        while not routing.IsEnd(idx3):
            next_idx = solution.Value(routing.NextVar(idx3))
            ni = manager.IndexToNode(idx3)
            nj = manager.IndexToNode(next_idx)
            sure_dk += data["duration_matrix"][ni][nj]
            if ni >= NUM_DEPOTS:
                sure_dk += data["service_time"]
            idx3 = next_idx
        sure_dk = round(sure_dk)
        asim    = "EVET" if sure_dk > 540 else "HAYIR"

        bg = C_SATIR1 if toplam_arac % 2 == 1 else C_SATIR2
        degerler = [
            f"Araç {v + 1}",
            f"Depo {baslangic_depo}",
            f"Depo {bitis_depo}",
            len(magaza_listesi),
            yuk,
            doluluk,
            round(mesafe_km, 1),
            rota_str,
            sure_dk,
            asim,
        ]
        ws_ozet.row_dimensions[satir].height = 18
        for col, val in enumerate(degerler, start=1):
            c = ws_ozet.cell(row=satir, column=col, value=val)
            c.font      = normal_font
            c.fill      = _doldur(bg)
            c.alignment = orta if col != 8 else sol
            c.border    = _kenar()

        ozet_satirlar.append({
            "arac_no":        v + 1,
            "baslangic":      baslangic_depo,
            "bitis":          bitis_depo,
            "magaza_listesi": magaza_listesi,
            "mesafe_km":      round(mesafe_km, 1),
            "yuk":            yuk,
            "doluluk":        doluluk,
            "sure_dk":        sure_dk,
            "asim":           asim,
        })
        satir += 1

    # Toplam satırı
    ws_ozet.row_dimensions[satir].height = 22
    toplam_vals = ["TOPLAM", "", "", "", "", "", round(toplam_mesafe, 1), "", "", ""]
    for col, val in enumerate(toplam_vals, start=1):
        c = ws_ozet.cell(row=satir, column=col, value=val)
        c.font      = toplam_font
        c.fill      = _doldur(C_TOPLAM_BG)
        c.alignment = orta
        c.border    = _kenar()
    ws_ozet.cell(row=satir, column=1).value = f"TOPLAM ({toplam_arac} Aktif Araç)"

    # Sütun genişlikleri
    genislikler = [14, 16, 16, 16, 20, 20, 22, 80, 18, 14]
    for i, g in enumerate(genislikler, start=1):
        ws_ozet.column_dimensions[get_column_letter(i)].width = g

    # ── Sayfa 2: Araç Detay Rotaları ───────────────────────────────────
    ws_det = wb.create_sheet("Araç Detay Rotaları")

    ws_det.merge_cells("A1:E1")
    c = ws_det["A1"]
    c.value     = "LC WİKİKİ — ARAÇ DETAY ROTA TABLOSU"
    c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=13)
    c.fill      = _doldur(C_BASLIK_BG)
    c.alignment = orta
    ws_det.row_dimensions[1].height = 28

    det_baslik = ["Araç No", "Sıra", "Düğüm No", "Tür", "Birikim Yük (koli)"]
    ws_det.row_dimensions[2].height = 18
    for col, bas in enumerate(det_baslik, start=1):
        c = ws_det.cell(row=2, column=col, value=bas)
        c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill      = _doldur(C_DEPO_BG)
        c.alignment = orta
        c.border    = _kenar()

    det_satir = 3
    for bilgi in ozet_satirlar:
        v_num = bilgi["arac_no"]
        # Depodan çıkış
        for col, val in enumerate([f"Araç {v_num}", 0, bilgi["baslangic"], "DEPO (Çıkış)", 0], start=1):
            c = ws_det.cell(row=det_satir, column=col, value=val)
            c.font      = Font(name="Calibri", bold=True, color=C_DEPO_FG, size=10)
            c.fill      = _doldur(C_DEPO_BG)
            c.alignment = orta
            c.border    = _kenar()
        ws_det.row_dimensions[det_satir].height = 16
        det_satir += 1

        # Mağaza ziyaretleri
        birikim = 0
        for sira, magaza_no in enumerate(bilgi["magaza_listesi"], start=1):
            birikim += STORE_DEMAND
            bg = C_MAĞAZA_BG if sira % 2 == 1 else C_SATIR2
            for col, val in enumerate([f"Araç {v_num}", sira, magaza_no, "Mağaza", birikim], start=1):
                c = ws_det.cell(row=det_satir, column=col, value=val)
                c.font      = normal_font
                c.fill      = _doldur(bg)
                c.alignment = orta
                c.border    = _kenar()
            ws_det.row_dimensions[det_satir].height = 15
            det_satir += 1

        # Depoya dönüş
        for col, val in enumerate([f"Araç {v_num}", "-", bilgi["bitis"], "DEPO (Dönüş)", birikim], start=1):
            c = ws_det.cell(row=det_satir, column=col, value=val)
            c.font      = Font(name="Calibri", bold=True, color=C_DEPO_FG, size=10)
            c.fill      = _doldur(C_DEPO_BG)
            c.alignment = orta
            c.border    = _kenar()
        ws_det.row_dimensions[det_satir].height = 16
        det_satir += 2  # Araçlar arası boşluk

    for i, g in enumerate([12, 8, 14, 18, 22], start=1):
        ws_det.column_dimensions[get_column_letter(i)].width = g

    # ── Sayfa 3: Araç Performans Özeti ─────────────────────────────────
    ws_perf = wb.create_sheet("Performans Özeti")

    ws_perf.merge_cells("A1:D1")
    c = ws_perf["A1"]
    c.value     = "ARAÇ PERFORMANS ÖZETİ"
    c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=13)
    c.fill      = _doldur(C_BASLIK_BG)
    c.alignment = orta
    ws_perf.row_dimensions[1].height = 28

    perf_bas = ["Metrik", "Değer", "Birim", "Not"]
    ws_perf.row_dimensions[2].height = 18
    for col, bas in enumerate(perf_bas, start=1):
        c = ws_perf.cell(row=2, column=col, value=bas)
        c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill      = _doldur(C_DEPO_BG)
        c.alignment = orta
        c.border    = _kenar()

    toplam_magaza = sum(len(b["magaza_listesi"]) for b in ozet_satirlar)
    ort_mesafe    = round(toplam_mesafe / len(ozet_satirlar), 1) if ozet_satirlar else 0
    ort_doluluk   = round(sum(b["doluluk"] for b in ozet_satirlar) / len(ozet_satirlar), 1) if ozet_satirlar else 0
    min_mesafe    = min(b["mesafe_km"] for b in ozet_satirlar) if ozet_satirlar else 0
    max_mesafe    = max(b["mesafe_km"] for b in ozet_satirlar) if ozet_satirlar else 0
    ort_sure      = round(sum(b["sure_dk"] for b in ozet_satirlar) / len(ozet_satirlar), 1) if ozet_satirlar else 0
    maks_sure     = max(b["sure_dk"] for b in ozet_satirlar) if ozet_satirlar else 0
    altinda       = sum(1 for b in ozet_satirlar if b["sure_dk"] <= 540)
    ustunde       = sum(1 for b in ozet_satirlar if b["sure_dk"] > 540)

    perf_veriler = [
        ("Toplam Aktif Araç",           toplam_arac,             "adet",  ""),
        ("Toplam Filo Mesafesi",         round(toplam_mesafe, 1), "km",    ""),
        ("Ortalama Araç Mesafesi",       ort_mesafe,              "km",    ""),
        ("En Kısa Rota",                 min_mesafe,              "km",    ""),
        ("En Uzun Rota",                 max_mesafe,              "km",    ""),
        ("Ziyaret Edilen Mağaza",        toplam_magaza,           "adet",  "Toplam 198 mağaza"),
        ("Ortalama Kapasite Doluluk",    ort_doluluk,             "%",     ""),
        ("Zaman Limiti",                 TIME_LIMIT_SECONDS,      "sn",    "GLS algoritması"),
        ("Araç Kapasitesi",              VEHICLE_CAPACITY,        "koli",  ""),
        ("Mağaza Talebi",                STORE_DEMAND,            "koli",  "Mağaza başına"),
        ("540 dk Altında Rota Sayısı",   altinda,                 "adet",  "Seyahat + 15 dk hizmet"),
        ("540 dk Üstünde Rota Sayısı",   ustunde,                 "adet",  "Seyahat + 15 dk hizmet"),
        ("Ortalama Rota Süresi (dk)",    ort_sure,                "dk",    ""),
        ("Maksimum Rota Süresi (dk)",    maks_sure,               "dk",    ""),
    ]

    for p_idx, (metrik, deger, birim, not_) in enumerate(perf_veriler, start=3):
        bg = C_SATIR1 if p_idx % 2 == 1 else C_SATIR2
        ws_perf.row_dimensions[p_idx].height = 18
        for col, val in enumerate([metrik, deger, birim, not_], start=1):
            c = ws_perf.cell(row=p_idx, column=col, value=val)
            c.font      = normal_font
            c.fill      = _doldur(bg)
            c.alignment = orta
            c.border    = _kenar()

    for i, g in enumerate([30, 16, 10, 28], start=1):
        ws_perf.column_dimensions[get_column_letter(i)].width = g

    # ── Sayfa 4: Maliyet Analizi ────────────────────────────────────────
    ws_mal = wb.create_sheet("Maliyet Analizi")

    # Maliyet hesaplamaları
    mal_yakit          = round(toplam_mesafe * YAKIT_TUKETIM_LT_KM * YAKIT_FIYATI_TL)
    mal_sabit          = toplam_arac * ARAC_GUNLUK_SABIT_TL
    mal_personel       = toplam_arac * SURUCU_GUNLUK_TL
    mal_inhouse        = mal_yakit + mal_sabit + mal_personel
    mal_tam_out        = TOPLAM_MAGAZA_SAYISI * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    mal_hibrit_out     = OUTSOURCE_MAGAZA_SAYISI * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    mal_hibrit_toplam  = mal_inhouse + mal_hibrit_out
    mal_tasarruf_gun   = mal_tam_out - mal_hibrit_toplam
    mal_tasarruf_yil   = mal_tasarruf_gun * 300
    mal_tasarruf_oran  = (mal_tasarruf_gun / mal_tam_out) * 100

    C_BOLUM_BG  = "2E4057"   # koyu lacivert-gri (bölüm başlıkları için)
    C_KAZANC_BG = "E2EFDA"   # açık yeşil (tasarruf satırları)
    C_KAZANC_FG = "375623"   # koyu yeşil yazı

    # Başlık
    ws_mal.merge_cells("A1:D1")
    c = ws_mal["A1"]
    c.value     = "LC WİKİKİ — MALİYET ANALİZİ (Günlük Bazda)"
    c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=14)
    c.fill      = _doldur(C_BASLIK_BG)
    c.alignment = orta
    ws_mal.row_dimensions[1].height = 32

    # Sütun başlıkları
    mal_bas = ["Kalem", "Tutar (TL)", "Birim", "Not"]
    ws_mal.row_dimensions[2].height = 18
    for col, bas in enumerate(mal_bas, start=1):
        c = ws_mal.cell(row=2, column=col, value=bas)
        c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill      = _doldur(C_DEPO_BG)
        c.alignment = orta
        c.border    = _kenar()

    def _bolum_baslik(ws, row, metin, renk=C_DEPO_BG):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=metin)
        c.font      = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill      = _doldur(renk)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _kenar()
        ws.row_dimensions[row].height = 18

    def _mal_satir(ws, row, kalem, tutar, birim, not_, bg=C_SATIR1, font=None):
        degerler = [kalem, tutar, birim, not_]
        ws.row_dimensions[row].height = 18
        for col, val in enumerate(degerler, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = font if font else Font(name="Calibri", size=10)
            c.fill      = _doldur(bg)
            c.alignment = orta if col != 1 else Alignment(horizontal="left", vertical="center")
            c.border    = _kenar()
            if col == 2 and isinstance(val, (int, float)):
                c.number_format = '#,##0'

    # Bölüm 1
    _bolum_baslik(ws_mal, 3, "  In-house Maliyet Kalemleri")
    _mal_satir(ws_mal, 4,  "Yakıt Maliyeti",
               mal_yakit,   "TL/gün",
               f"{toplam_mesafe:.0f} km × {YAKIT_TUKETIM_LT_KM} lt × {YAKIT_FIYATI_TL} TL")
    _mal_satir(ws_mal, 5,  "Araç Sabit Maliyeti",
               mal_sabit,   "TL/gün",
               f"{toplam_arac} araç × {ARAC_GUNLUK_SABIT_TL:,} TL", bg=C_SATIR2)
    _mal_satir(ws_mal, 6,  "Sürücü Personel Maliyeti",
               mal_personel, "TL/gün",
               f"{toplam_arac} araç × {SURUCU_GUNLUK_TL:,} TL")
    _mal_satir(ws_mal, 7,  "In-house Günlük Toplam",
               mal_inhouse, "TL/gün", "",
               bg=C_TOPLAM_BG,
               font=Font(name="Calibri", bold=True, color=C_TOPLAM_FG, size=10))

    # Bölüm 2
    _bolum_baslik(ws_mal, 8, "  Outsource (3PL) Maliyet")
    _mal_satir(ws_mal, 9,  "Tam Outsource Maliyeti",
               mal_tam_out, "TL/gün",
               f"{TOPLAM_MAGAZA_SAYISI} mağaza × {STORE_DEMAND} koli × {OUTSOURCE_KOLI_FIYAT_TL} TL")
    _mal_satir(ws_mal, 10, "Hibrit Outsource Kısmı",
               mal_hibrit_out, "TL/gün",
               f"{OUTSOURCE_MAGAZA_SAYISI} mağaza × {STORE_DEMAND} koli × {OUTSOURCE_KOLI_FIYAT_TL} TL",
               bg=C_SATIR2)

    # Bölüm 3
    _bolum_baslik(ws_mal, 11, "  Karşılaştırma", renk=C_BOLUM_BG)
    _mal_satir(ws_mal, 12, "Mevcut Sistem (Tam 3PL)",
               mal_tam_out, "TL/gün", f"{TOPLAM_MAGAZA_SAYISI} mağaza")
    _mal_satir(ws_mal, 13, "Önerilen Hibrit Sistem",
               mal_hibrit_toplam, "TL/gün",
               f"{INHOUSE_MAGAZA_SAYISI} in-house + {OUTSOURCE_MAGAZA_SAYISI} outsource",
               bg=C_SATIR2)
    _mal_satir(ws_mal, 14, "Günlük Tasarruf",
               round(mal_tasarruf_gun), "TL/gün", "Tam 3PL - Hibrit",
               bg=C_KAZANC_BG,
               font=Font(name="Calibri", bold=True, color=C_KAZANC_FG, size=10))
    _mal_satir(ws_mal, 15, "Yıllık Tasarruf (300 iş günü)",
               round(mal_tasarruf_yil), "TL/yıl", "Günlük × 300",
               bg=C_KAZANC_BG,
               font=Font(name="Calibri", bold=True, color=C_KAZANC_FG, size=11))
    _mal_satir(ws_mal, 16, "Tasarruf Oranı",
               f"%{mal_tasarruf_oran:.1f}", "", "(Tam 3PL baz alınarak)",
               bg=C_KAZANC_BG,
               font=Font(name="Calibri", bold=True, color=C_KAZANC_FG, size=10))

    for i, g in enumerate([38, 20, 12, 40], start=1):
        ws_mal.column_dimensions[get_column_letter(i)].width = g

    # Kaydet
    wb.save(cikti_yolu)
    print(f"\n  Excel raporu kaydedildi: {cikti_yolu}")


# ─────────────────────────────────────────────
# 6. ANA PROGRAM
# ─────────────────────────────────────────────

def main():
    print()
    print("=" * 70)
    print("  LC WİKİKİ — MDCVRP OPTİMİZASYON MOTORU (OR-Tools)")
    print("=" * 70)

    print("\n[1/3] Veri yükleniyor...")
    mesafe_matrisi = yukle_mesafe_matrisi(EXCEL_PATH, SHEET_NAME)
    sure_matrisi   = yukle_sure_matrisi(EXCEL_PATH, SHEET_NAME_SURE)

    print("\n[2/3] Model oluşturuluyor...")
    data = veri_modeli_olustur(mesafe_matrisi, sure_matrisi)
    print(f"  Toplam lokasyon : {data['num_locations']}")
    print(f"  Depo sayısı     : {NUM_DEPOTS}")
    print(f"  Mağaza sayısı   : {data['num_locations'] - NUM_DEPOTS}")
    print(f"  Araç sayısı     : {NUM_VEHICLES}")
    print(f"  Araç kapasitesi : {VEHICLE_CAPACITY} koli")
    print(f"  Mağaza talebi   : {STORE_DEMAND} koli")
    print(f"  Araç başlangıçları: {VEHICLE_STARTS}")

    print("\n[3/3] Optimizasyon çalışıyor...")
    manager, routing, solution = coz(data)

    rapor_yazdir(manager, routing, solution, data)

    tarih_damga = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_yolu  = rf"C:\Users\Zahid Sami\gams_output\LCW_MDCVRP_Sonuc_{tarih_damga}.xlsx"
    print("[4/4] Excel raporu oluşturuluyor...")
    excel_export(manager, routing, solution, data, excel_yolu)


if __name__ == "__main__":
    main()
