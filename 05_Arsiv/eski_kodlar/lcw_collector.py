#!/usr/bin/env python3
"""
LCW-Collector v2.0
LC Waikiki Pilot İl Mağaza Lokasyon Toplayıcısı
GÜNCELLEME: İstanbul ve Bölgesel Kümeleme eklendi, Sayfa limitleri artırıldı.
"""
from __future__ import annotations

import json
import time
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─── YAPILANDIRMA ────────────────────────────────────────────────────────────
SERPAPI_KEY = "23b78dd28203569a8abfd5a2d5151f6329c1b0b5f80f3a11d37e4c927aa5d6eb"
SERPAPI_ENDPOINT = "https://serpapi.com/search"

# İSTANBUL LİSTEYE EKLENDİ
TARGET_CITIES = [
    "İstanbul", "Van", "Erzurum", "Ankara", "Antalya", "Afyonkarahisar",
    "Adana", "Kayseri", "Diyarbakır", "Gaziantep",
    "Bursa", "Trabzon", "Samsun", "İzmir",
]

# İSTANBUL VARYANTLARI EKLENDİ
CITY_ALIASES: dict[str, list[str]] = {
    "İstanbul":   ["İstanbul", "Istanbul", "ist"],
    "Van":        ["Van"],
    "Erzurum":    ["Erzurum"],
    "Ankara":     ["Ankara"],
    "Antalya":    ["Antalya"],
    "Afyon":      ["Afyon", "Afyonkarahisar"],
    "Adana":      ["Adana"],
    "Kayseri":    ["Kayseri"],
    "Diyarbakır": ["Diyarbakır", "Diyarbakir"],
    "Gaziantep":  ["Gaziantep"],
    "Bursa":      ["Bursa"],
    "Trabzon":    ["Trabzon"],
    "Samsun":     ["Samsun"],
    "İzmir":      ["İzmir", "Izmir"],
}

# İstanbul Yakaları İçin İlçe Veritabanı
ANADOLU_YAKASI = [
    "adalar", "ataşehir", "beykoz", "çekmeköy", "kadıköy", "kartal", 
    "maltepe", "pendik", "sancaktepe", "sultanbeyli", "şile", "tuzla", 
    "ümraniye", "üsküdar"
]

MAX_PAGES_PER_CITY = 12     # İstanbul için sayfa limiti 5'ten 12'ye çıkarıldı (Max 240 Mağaza)
RESULTS_PER_PAGE   = 20     
REQUEST_DELAY      = 1.5    
MAX_RETRIES        = 3      
RATE_LIMIT_WAIT    = 60     

CACHE_DIR   = Path("cache/serpapi_raw")
LOG_DIR     = Path("logs")
OUTPUT_FILE = "LCW_Pilot_Lokasyonlar.xlsx"

TURKEY_LAT = (36.0, 42.5)
TURKEY_LON = (26.0, 45.0)

# ─── LOGGING KURULUMU ────────────────────────────────────────────────────────
def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file  = LOG_DIR / f"run_{timestamp}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger("lcw-collector")

# ─── ÖNBELLEK KATMANI ────────────────────────────────────────────────────────
class CacheLayer:
    def __init__(self, cache_dir: Path):
        self.cache_dir = cache_dir
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _path(self, city: str, page: int) -> Path:
        safe = city.replace("İ", "I").replace("ı", "i").replace("ğ", "g").replace("ş", "s")
        return self.cache_dir / f"{safe}_page{page}.json"

    def get(self, city: str, page: int) -> Optional[dict]:
        p = self._path(city, page)
        if p.exists():
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        return None

    def set(self, city: str, page: int, data: dict) -> None:
        p = self._path(city, page)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def exists(self, city: str, page: int) -> bool:
        return self._path(city, page).exists()

# ─── SERPAPI İSTEMCİSİ ───────────────────────────────────────────────────────
class SerpApiClient:
    def __init__(self, api_key: str, cache: CacheLayer, logger: logging.Logger):
        self.api_key     = api_key
        self.cache       = cache
        self.logger      = logger
        self.total_calls = 0

    def _request(self, params: dict, attempt: int = 1) -> Optional[dict]:
        try:
            resp = requests.get(SERPAPI_ENDPOINT, params=params, timeout=30)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code == 429:
                if attempt <= MAX_RETRIES:
                    self.logger.warning(f"Rate limit aşıldı. {RATE_LIMIT_WAIT}s bekleniyor... (Deneme {attempt}/{MAX_RETRIES})")
                    time.sleep(RATE_LIMIT_WAIT)
                    return self._request(params, attempt + 1)
                return None
            if resp.status_code in (401, 403):
                self.logger.critical("KİMLİK DOĞRULAMA HATASI. API anahtarını kontrol edin!")
                raise SystemExit(1)
            if resp.status_code >= 500:
                if attempt <= 2:
                    time.sleep(30)
                    return self._request(params, attempt + 1)
                return None
            return None
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            if attempt <= MAX_RETRIES:
                time.sleep(5 * (2 ** (attempt - 1)))
                return self._request(params, attempt + 1)
            return None

    def search_city(self, city: str) -> list[dict]:
        all_results: list[dict] = []
        for page in range(MAX_PAGES_PER_CITY):
            start = page * RESULTS_PER_PAGE
            cached = self.cache.get(city, page)
            if cached is not None:
                self.logger.info(f"  [ÖNBELLEK] {city} sayfa {page} → diskten yüklendi")
                results = cached.get("local_results", [])
                all_results.extend(results)
                if len(results) < RESULTS_PER_PAGE:
                    break
                continue

            params = {
                "engine":  "google_maps",
                "q":       f"LC Waikiki {city}",
                "type":    "search",
                "hl":      "tr",
                "gl":      "tr",
                "start":   start,
                "api_key": self.api_key,
            }

            self.logger.info(f"  [API] {city} sayfa {page} çekiliyor...")
            data = self._request(params)
            self.total_calls += 1

            if data is None: break
            self.cache.set(city, page, data)
            results = data.get("local_results", [])
            all_results.extend(results)

            if len(results) < RESULTS_PER_PAGE:
                break
            time.sleep(REQUEST_DELAY)
        return all_results

# ─── ALAN ÇIKARICISI VE BÖLGESEL KÜMELEME ────────────────────────────────────
class FieldExtractor:
    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def extract(self, raw: dict, query_city: str) -> Optional[dict]:
        try:
            name    = raw.get("title", "").strip()
            address = raw.get("address", "").strip()
            gps     = raw.get("gps_coordinates", {})
            lat     = gps.get("latitude")
            lon     = gps.get("longitude")

            if not name or not address or lat is None or lon is None:
                return None

            province, district = self._parse_address(address)
            province_final = province or query_city
            
            # Bölgesel Etiketleme (İstanbul Yakaları)
            bolge = province_final
            if province_final.lower() in ["istanbul", "i̇stanbul", "ist"]:
                dist_lower = district.lower()
                if any(anadolu_ilce in dist_lower for anadolu_ilce in ANADOLU_YAKASI):
                    bolge = "İstanbul - Anadolu"
                else:
                    bolge = "İstanbul - Avrupa"

            return {
                "Mağaza Adı":  name,
                "Bölge":       bolge,
                "İl":          province_final,
                "İlçe":        district,
                "Tam Adres":   address,
                "Enlem":       round(float(lat), 6),
                "Boylam":      round(float(lon), 6),
                "_query_city": query_city,
            }
        except Exception as exc:
            self.logger.error(f"Alan çıkarımı hatası: {exc}")
            return None

    def _parse_address(self, address: str) -> tuple[str, str]:
        province = district = ""
        if not address: return province, district
        last = address.split(",")[-1].strip()
        tokens      = last.split()
        no_postal   = [t for t in tokens if not (t.isdigit() and len(t) == 5)]
        location    = " ".join(no_postal).strip()
        if "/" in location:
            parts    = location.split("/")
            district = parts[0].strip()
            province = parts[-1].strip()
        else:
            province = location
        return province, district

# ─── COĞRAFİ FİLTRE ──────────────────────────────────────────────────────────
class GeoFilter:
    def __init__(self, logger: logging.Logger):
        self.logger  = logger
        self.removed = []
        self.allowed: set[str] = set()
        for aliases in CITY_ALIASES.values():
            for alias in aliases:
                self.allowed.add(alias.lower())

    def _is_allowed(self, record: dict) -> bool:
        province   = record.get("İl", "").strip().lower()
        query_city = record.get("_query_city", "").strip().lower()
        if province in self.allowed or query_city in self.allowed:
            return True
        return False

    def filter(self, records: list[dict]) -> list[dict]:
        passed = []
        for rec in records:
            if self._is_allowed(rec):
                passed.append(rec)
            else:
                self.removed.append(rec)
        return passed

    def save_log(self, log_dir: Path) -> None:
        if self.removed:
            pd.DataFrame(self.removed).to_csv(log_dir / "filtered_out_log.csv", index=False, encoding="utf-8-sig")

# ─── TEKİLLEŞTİRME MOTORU ───────────────────────────────────────────────────
class DeduplicationEngine:
    def __init__(self, logger: logging.Logger):
        self.logger     = logger
        self.duplicates = []

    def _key(self, rec: dict) -> str:
        return f"{rec.get('Mağaza Adı', '').lower().strip()}|{rec.get('Tam Adres', '').lower().strip()}"

    def deduplicate(self, records: list[dict]) -> list[dict]:
        seen:   dict[str, dict] = {}
        unique: list[dict]      = []
        for rec in records:
            k = self._key(rec)
            if k not in seen:
                seen[k] = rec
                unique.append(rec)
            else:
                self.duplicates.append(rec)
        return unique

    def save_log(self, log_dir: Path) -> None:
        if self.duplicates:
            pd.DataFrame(self.duplicates).to_csv(log_dir / "duplicate_log.csv", index=False, encoding="utf-8-sig")

def validate_coordinates(records: list[dict], logger: logging.Logger) -> list[dict]:
    valid = []
    for rec in records:
        lat, lon = rec.get("Enlem"), rec.get("Boylam")
        if lat and lon and (TURKEY_LAT[0] <= lat <= TURKEY_LAT[1]) and (TURKEY_LON[0] <= lon <= TURKEY_LON[1]):
            valid.append(rec)
    return valid

# ─── EXCEL DIŞA AKTARICISI ───────────────────────────────────────────────────
class ExcelExporter:
    HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HDR_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    ALT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def _style_header(self, ws, row: int = 1) -> None:
        for cell in ws[row]:
            cell.fill, cell.font, cell.alignment = self.HDR_FILL, self.HDR_FONT, Alignment(horizontal="center", vertical="center")

    def _sheet_main(self, wb: Workbook, records: list[dict]) -> None:
        ws = wb.active
        ws.title = "Mağazalar"
        # BÖLGE SÜTUNU EKLENDİ
        columns    = ["Mağaza Adı", "Bölge", "İl", "İlçe", "Tam Adres", "Enlem", "Boylam"]
        col_widths = [35,           20,      15,   20,     60,          12,      12]

        ws.append(columns)
        self._style_header(ws)

        for row_i, rec in enumerate(records, start=2):
            ws.append([rec.get(c, "") for c in columns])
            if row_i % 2 == 0:
                for cell in ws[row_i]: cell.fill = self.ALT_FILL

        for col_i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = width
        ws.freeze_panes = "A2"

    def _sheet_summary(self, wb: Workbook, records: list[dict], api_calls: int) -> None:
        ws = wb.create_sheet("Özet İstatistikler")
        ws.append(["Metrik", "Değer"])
        self._style_header(ws)
        ws.append(["Toplam Mağaza", len(records)])
        ws.append(["API Çağrısı", api_calls])

    def export(self, records: list[dict], output_path: str, api_calls: int) -> None:
        wb = Workbook()
        self._sheet_main(wb, records)
        self._sheet_summary(wb, records, api_calls)
        wb.save(output_path)
        self.logger.info(f"Excel dosyası kaydedildi: {output_path}")

# ─── ANA ORKESTRATÖR ─────────────────────────────────────────────────────────
def main() -> None:
    logger = setup_logging()
    cache  = CacheLayer(CACHE_DIR)
    client = SerpApiClient(SERPAPI_KEY, cache, logger)
    extractor    = FieldExtractor(logger)
    geo_filter   = GeoFilter(logger)
    deduplicator = DeduplicationEngine(logger)

    raw_extracted = []
    for city in TARGET_CITIES:
        for result in client.search_city(city):
            rec = extractor.extract(result, city)
            if rec: raw_extracted.append(rec)

    filtered_records = geo_filter.filter(raw_extracted)
    validated        = validate_coordinates(filtered_records, logger)
    unique_records   = deduplicator.deduplicate(validated)

    for rec in unique_records: rec.pop("_query_city", None)

    ExcelExporter(logger).export(unique_records, OUTPUT_FILE, client.total_calls)
    logger.info(f"İŞLEM TAMAMLANDI! Toplam {len(unique_records)} mağaza {OUTPUT_FILE} dosyasına kaydedildi.")

if __name__ == "__main__":
    main()