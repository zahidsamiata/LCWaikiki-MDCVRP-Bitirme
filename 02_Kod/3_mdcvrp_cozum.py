"""
LC Waikiki — MDCVRP 200 km Eşiği
6 Gerçek Depo + ≤200 km altındaki mağazaların alt matrisi
Google OR-Tools: SAVINGS başlangıç + GLS meta-sezgisel
"""

import math
import time
import datetime
import numpy as np
import pandas as pd
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ortools.constraint_solver import routing_enums_pb2, pywrapcp
from pathlib import Path

# ─── DOSYA YOLLARI ────────────────────────────────────────────────────────────
_PROJE_KOK  = Path(__file__).resolve().parent.parent
LOK_PATH    = _PROJE_KOK / "01_Veri" / "LCW_Lokasyonlar.xlsx"
MATRIX_PATH = _PROJE_KOK / "01_Veri" / "LCW_Mesafe_Sure_Matrisi.xlsx"
OUTPUT_DIR  = _PROJE_KOK / "04_Sonuclar"

ESIK_KM             = 200
NUM_DEPOTS          = 6

# ─── ARAÇ DAĞILIMI ────────────────────────────────────────────────────────────
# Depo indeksleri: LM=0, Eroğlu=1, Silivri=2, Hadımköy=3, Yalova=4, Aksaray=5
#
# NOT (kod incelemesi sonrasi eklendi): Bu dagilim, once "her magaza en yakin
# depoya atanir" varsayimiyla depo basi magaza sayisi (depot_count, asagida
# hazirla_alt_problem() icinde hesaplanir) incelenerek, sonra toplam talep /
# arac kapasitesi oranindan (asagidaki min_arac_ihtiyaci) yola cikilarak,
# son olarak 540 dk/gun rota suresi ustunde kalan araclara ek arac eklenerek
# elle belirlenmistir. Silivri (depo 2) bilercek atlanmamistir: 200 km esigi
# altinda Silivri'ye en yakin hicbir magaza yoktur (depot_count[2] == 0),
# dolayisiyla bu depoya arac atanmasina gerek kalmamistir — asagidaki
# hazirla_alt_problem() ciktisinda bu dogrulanabilir.
#
# Bu manuel adim, modelin "genel" formulasyonunu (Bolum 3.4) degil, bu ozel
# 204 magazalik pilot orneginin filo boyutlandirma adimini ilgilendirir;
# farkli bir veri setinde MANUAL_VEHICLE_STARTS/ENDS yeniden belirlenmelidir.
MANUAL_VEHICLE_STARTS = [0, 1, 3, 4, 4, 5]
MANUAL_VEHICLE_ENDS   = [0, 1, 3, 4, 4, 5]

# ─── MALİYET PARAMETRELERİ ────────────────────────────────────────────────────
STORE_DEMAND            = 150
VEHICLE_CAPACITY        = 3000
TIME_LIMIT_SECONDS      = 60
SERVICE_TIME_DK         = 15
YAKIT_FIYATI_TL         = 43.0
YAKIT_TUKETIM_LT_KM     = 0.35
ARAC_GUNLUK_SABIT_TL    = 2500
SURUCU_GUNLUK_TL        = 1500
OUTSOURCE_KOLI_FIYAT_TL = 3.75
PILOT_MAGAZA_SAYISI     = 204   # karşılaştırma bazı (pilot kapsam)
TOPLAM_MAGAZA_HAM       = 533   # kapsam-uyumsuz ham referans (dipnot için)


# ─────────────────────────────────────────────────────────────────────────────
# 1. VERİ YÜKLEME VE ALT MATRİS HAZIRLIK
# ─────────────────────────────────────────────────────────────────────────────

def hazirla_alt_problem():
    print("  Lokasyon dosyası okunuyor...")
    df_lok = pd.read_excel(LOK_PATH)
    depo_adlari = list(df_lok["Mağaza Adı"][:NUM_DEPOTS])
    print(f"  Depolar: {depo_adlari}")

    print("  Mesafe matrisi okunuyor (210×210)...")
    df_km = pd.read_excel(MATRIX_PATH, sheet_name="Mesafe (km)", index_col=0)
    df_dk = pd.read_excel(MATRIX_PATH, sheet_name="Süre (dk)",   index_col=0)
    print(f"  Matris boyutu: {df_km.shape}")

    full_km = df_km.fillna(0).values   # 210×210, float
    full_dk = df_dk.fillna(0).values   # 210×210, float

    # ── Her mağaza için en yakın depo ve mesafeyi bul ──────────────────────
    depo_idx   = list(range(NUM_DEPOTS))
    magaza_idx = list(range(NUM_DEPOTS, full_km.shape[0]))  # [6..209]

    en_yakin_mesafe = np.array([
        min(full_km[m][d] for d in depo_idx) for m in magaza_idx
    ])
    en_yakin_depo = np.array([
        depo_idx[np.argmin([full_km[m][d] for d in depo_idx])]
        for m in magaza_idx
    ])

    # ── 200 km eşiği ──────────────────────────────────────────────────────
    inhouse_mask    = en_yakin_mesafe <= ESIK_KM
    inhouse_local   = np.where(inhouse_mask)[0]                   # yerel [0..203]
    inhouse_global  = [magaza_idx[i] for i in inhouse_local]      # global [6..209]
    outsource_count = int((~inhouse_mask).sum())
    inhouse_count   = int(inhouse_mask.sum())

    print(f"\n  ≤{ESIK_KM} km eşiği → In-house: {inhouse_count}, Outsource: {outsource_count}")

    # ── Alt matris indeksleri ──────────────────────────────────────────────
    sub_global = list(range(NUM_DEPOTS)) + inhouse_global   # [0..5, g6, g7, ...]
    n_sub = len(sub_global)

    # km → metre → int (OR-Tools için)
    sub_dist = [
        [int(full_km[sub_global[i]][sub_global[j]] * 1000) for j in range(n_sub)]
        for i in range(n_sub)
    ]
    # süre: dakika → int (yuvarla)
    sub_dur = [
        [int(round(full_dk[sub_global[i]][sub_global[j]])) for j in range(n_sub)]
        for i in range(n_sub)
    ]

    # ── Depo başına mağaza sayısı (200 km içinde) ─────────────────────────
    depot_count = Counter(en_yakin_depo[inhouse_mask])  # {depo_idx: n_stores}
    for d in depo_idx:
        depot_count.setdefault(d, 0)

    print("\n  Depo başı mağaza dağılımı (≤200 km):")
    for d in depo_idx:
        print(f"    Depo {d} ({depo_adlari[d]}): {depot_count[d]} mağaza")

    # DOGRULAMA: MANUAL_VEHICLE_STARTS icinde olmayan depolarin gercekten
    # 0 magazali oldugunu kontrol et (aksi halde o depoya en yakin magazalar
    # baska depodan hizmet alir ve bu durum ayrica raporlanmalidir).
    for d in depo_idx:
        if d not in MANUAL_VEHICLE_STARTS and depot_count[d] > 0:
            print(f"    [UYARI] Depo {d} ({depo_adlari[d]}) icin arac atanmamis "
                  f"ancak {depot_count[d]} magaza bu depoya en yakin. "
                  f"Bu magazalar baska depodan araclarla hizmet alacak.")

    # ── Araç sayısı ve başlangıç noktaları (manuel dağılım) ──────────────────
    vehicle_starts = MANUAL_VEHICLE_STARTS
    vehicle_ends   = MANUAL_VEHICLE_ENDS
    num_vehicles   = len(vehicle_starts)

    active_depots = sorted(set(vehicle_starts))
    allocation    = Counter(vehicle_starts)

    print(f"\n  Araç sayısı    : {num_vehicles}")
    print(f"  Aktif depolar  : {active_depots}")
    print(f"  Araç dağılımı  : {dict(sorted(allocation.items()))}")
    print(f"  Starts/Ends    : {vehicle_starts}")

    return {
        "sub_dist":        sub_dist,
        "sub_dur":         sub_dur,
        "sub_global":      sub_global,      # local → global indeks
        "inhouse_global":  inhouse_global,
        "inhouse_count":   inhouse_count,
        "outsource_count": outsource_count,
        "num_vehicles":    num_vehicles,
        "vehicle_starts":  vehicle_starts,
        "vehicle_ends":    vehicle_ends,
        "depot_count":     depot_count,
        "depo_adlari":     depo_adlari,
        "allocation":      allocation,
        "active_depots":   active_depots,
        "n_sub":           n_sub,
        "en_yakin_mesafe": en_yakin_mesafe,
        "en_yakin_depo":   en_yakin_depo,
        "inhouse_mask":    inhouse_mask,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 2. OR-TOOLS VERİ MODELİ
# ─────────────────────────────────────────────────────────────────────────────

def veri_modeli_olustur(prob: dict) -> dict:
    n = prob["n_sub"]
    demands = [0] * NUM_DEPOTS + [STORE_DEMAND] * (n - NUM_DEPOTS)

    return {
        "distance_matrix":    prob["sub_dist"],
        "duration_matrix":    prob["sub_dur"],
        "demands":            demands,
        "vehicle_capacities": [VEHICLE_CAPACITY] * prob["num_vehicles"],
        "num_vehicles":       prob["num_vehicles"],
        "starts":             prob["vehicle_starts"],
        "ends":               prob["vehicle_ends"],
        "num_locations":      n,
        "service_time":       SERVICE_TIME_DK,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 3. ÇÖZÜCÜ
# ─────────────────────────────────────────────────────────────────────────────

def coz(data: dict):
    manager = pywrapcp.RoutingIndexManager(
        data["num_locations"],
        data["num_vehicles"],
        data["starts"],
        data["ends"],
    )
    routing = pywrapcp.RoutingModel(manager)

    def distance_callback(from_idx, to_idx):
        return data["distance_matrix"][manager.IndexToNode(from_idx)][manager.IndexToNode(to_idx)]

    transit_cb = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_cb)

    def demand_callback(idx):
        return data["demands"][manager.IndexToNode(idx)]

    demand_cb = routing.RegisterUnaryTransitCallback(demand_callback)
    routing.AddDimensionWithVehicleCapacity(
        demand_cb, 0, data["vehicle_capacities"], True, "Capacity"
    )

    search_params = pywrapcp.DefaultRoutingSearchParameters()
    search_params.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.SAVINGS
    )
    search_params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    )
    search_params.time_limit.seconds = TIME_LIMIT_SECONDS
    search_params.log_search = False

    print(f"\n  Optimizasyon başlatılıyor ({TIME_LIMIT_SECONDS} sn)...")
    t0 = time.time()
    solution = routing.SolveWithParameters(search_params)
    elapsed = time.time() - t0
    print(f"  Tamamlandı: {elapsed:.1f} sn")

    return manager, routing, solution


# ─────────────────────────────────────────────────────────────────────────────
# 4. KONSOL RAPORU
# ─────────────────────────────────────────────────────────────────────────────

SEP  = "─" * 72
SEP2 = "═" * 72

def rapor_yazdir(manager, routing, solution, data, prob):
    if not solution:
        print("\n  [HATA] Çözüm bulunamadı!")
        return None

    cap_dim = routing.GetDimensionOrDie("Capacity")
    sub_global = prob["sub_global"]
    depo_adlari = prob["depo_adlari"]

    print()
    print(SEP2)
    print("  LC WİKİKİ — MDCVRP 200 KM EŞİĞİ — FİLO RAPORU")
    print(SEP2)

    toplam_mesafe = 0.0
    toplam_yuklu  = 0
    ozet_satirlar = []

    for v in range(data["num_vehicles"]):
        idx = routing.Start(v)
        baslangic_local = manager.IndexToNode(idx)

        rota_local = []
        while not routing.IsEnd(idx):
            rota_local.append(manager.IndexToNode(idx))
            idx = solution.Value(routing.NextVar(idx))
        bitis_local = manager.IndexToNode(idx)

        magaza_local = [n for n in rota_local if n >= NUM_DEPOTS]
        if not magaza_local:
            continue

        toplam_yuklu += 1

        mesafe_m = 0
        idx2 = routing.Start(v)
        while not routing.IsEnd(idx2):
            ni = solution.Value(routing.NextVar(idx2))
            mesafe_m += data["distance_matrix"][manager.IndexToNode(idx2)][manager.IndexToNode(ni)]
            idx2 = ni
        mesafe_km = mesafe_m / 1000
        toplam_mesafe += mesafe_km

        yuk = solution.Value(cap_dim.CumulVar(routing.End(v)))
        doluluk = round(yuk / VEHICLE_CAPACITY * 100, 1)

        sure_dk = 0
        idx3 = routing.Start(v)
        while not routing.IsEnd(idx3):
            ni3 = solution.Value(routing.NextVar(idx3))
            n_from = manager.IndexToNode(idx3)
            n_to   = manager.IndexToNode(ni3)
            sure_dk += data["duration_matrix"][n_from][n_to]
            if n_from >= NUM_DEPOTS:
                sure_dk += SERVICE_TIME_DK
            idx3 = ni3
        sure_dk = round(sure_dk)

        # Global indeksler (kullanıcının 210×210 matrisindeki satır numaraları)
        magaza_global = [sub_global[n] for n in magaza_local]
        rota_str = " → ".join(str(g) for g in magaza_global)

        print(f"\n  ARAÇ {v+1:>2d} | Depo {baslangic_local} ({depo_adlari[baslangic_local]})")
        print(f"  {SEP}")
        print(f"  Mağaza Sayısı  : {len(magaza_local)}")
        print(f"  Yük            : {yuk} / {VEHICLE_CAPACITY} koli  ({doluluk}%)")
        print(f"  Mesafe         : {mesafe_km:.1f} km")
        print(f"  Süre           : {sure_dk} dk  {'[AŞIM >540dk]' if sure_dk > 540 else ''}")
        satirlar = [magaza_global[i:i+10] for i in range(0, len(magaza_global), 10)]
        for k, sat in enumerate(satirlar):
            lbl = "Rota (global)" if k == 0 else ""
            print(f"  {lbl:<15}: " + " → ".join(str(g) for g in sat))

        ozet_satirlar.append({
            "arac_no":        v + 1,
            "baslangic":      baslangic_local,
            "bitis":          bitis_local,
            "magaza_local":   magaza_local,
            "magaza_global":  magaza_global,
            "mesafe_km":      round(mesafe_km, 1),
            "yuk":            yuk,
            "doluluk":        doluluk,
            "sure_dk":        sure_dk,
            "asim":           "EVET" if sure_dk > 540 else "HAYIR",
        })

    print()
    print(SEP2)
    print(f"  TOPLAM AKTİF ARAÇ   : {toplam_yuklu} / {data['num_vehicles']}")
    print(f"  TOPLAM FİLO MESAFESİ: {toplam_mesafe:.1f} km")
    print(f"  OBJEKTİF (metre)    : {solution.ObjectiveValue():,}")
    print(SEP2)

    # ── Maliyet Analizi ──────────────────────────────────────────────────
    inhouse_count   = prob["inhouse_count"]
    outsource_count = prob["outsource_count"]

    yakit          = toplam_mesafe * YAKIT_TUKETIM_LT_KM * YAKIT_FIYATI_TL
    sabit          = toplam_yuklu * ARAC_GUNLUK_SABIT_TL
    personel       = toplam_yuklu * SURUCU_GUNLUK_TL
    inhouse_toplam = yakit + sabit + personel
    pilot_outsource   = PILOT_MAGAZA_SAYISI * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    tam_outsource_ham = TOPLAM_MAGAZA_HAM * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    hibrit_out        = outsource_count * STORE_DEMAND * OUTSOURCE_KOLI_FIYAT_TL
    hibrit_toplam     = inhouse_toplam + hibrit_out
    tasarruf_gun      = pilot_outsource - hibrit_toplam
    tasarruf_yil      = tasarruf_gun * 300
    tasarruf_oran     = (tasarruf_gun / pilot_outsource) * 100

    print()
    print(SEP2)
    print("  MALİYET ANALİZİ")
    print(SEP2)
    print(f"  In-house mağaza     : {inhouse_count}")
    print(f"  Outsource mağaza    : {outsource_count}")
    print(f"  Aktif araç sayısı   : {toplam_yuklu}")
    print()
    print(f"  Yakıt               : {yakit:>12,.0f} TL/gün")
    print(f"  Araç sabit          : {sabit:>12,.0f} TL/gün")
    print(f"  Sürücü              : {personel:>12,.0f} TL/gün")
    print(f"  In-house toplam     : {inhouse_toplam:>12,.0f} TL/gün")
    print(f"  {SEP}")
    print(f"  Pilot tam out. (204): {pilot_outsource:>12,.0f} TL/gün")
    print(f"  Hibrit outsource    : {hibrit_out:>12,.0f} TL/gün")
    print(f"  Hibrit toplam       : {hibrit_toplam:>12,.0f} TL/gün")
    print(f"  [Ref-Ham] Tam (533) : {tam_outsource_ham:>12,.0f} TL/gün  (kapsam-uyumsuz)")
    print(f"  {SEP}")
    print(f"  Günlük tasarruf     : {tasarruf_gun:>+12,.0f} TL/gün")
    print(f"  Yıllık tasarruf     : {tasarruf_yil:>+12,.0f} TL/yıl")
    print(f"  Tasarruf oranı      : %{tasarruf_oran:>+.1f}")
    print(SEP2)

    maliyet = {
        "toplam_mesafe": toplam_mesafe,
        "toplam_yuklu":  toplam_yuklu,
        "yakit":         yakit,
        "sabit":         sabit,
        "personel":      personel,
        "inhouse_toplam":    inhouse_toplam,
        "pilot_outsource":   pilot_outsource,
        "tam_outsource_ham": tam_outsource_ham,
        "hibrit_out":        hibrit_out,
        "hibrit_toplam":     hibrit_toplam,
        "tasarruf_gun":      tasarruf_gun,
        "tasarruf_yil":      tasarruf_yil,
        "tasarruf_oran":     tasarruf_oran,
    }
    return ozet_satirlar, maliyet


# ─────────────────────────────────────────────────────────────────────────────
# 5. EXCEL EXPORT (5 sayfa)
# ─────────────────────────────────────────────────────────────────────────────

C_BASLIK_BG = "1F3864"
C_BASLIK_FG = "FFFFFF"
C_DEPO_BG   = "2E75B6"
C_DEPO_FG   = "FFFFFF"
C_TOPLAM_BG = "FFF2CC"
C_TOPLAM_FG = "7B3F00"
C_SATIR1    = "EBF3FB"
C_SATIR2    = "FFFFFF"
C_MAGAZA_BG = "E2EFDA"
C_KAZANC_BG = "E2EFDA"
C_KAZANC_FG = "375623"
C_KENARLIK  = "BDD7EE"
C_BOLUM_BG  = "2E4057"


def _fill(hex_kod):
    return PatternFill("solid", fgColor=hex_kod)

def _border():
    s = Side(style="thin", color=C_KENARLIK)
    return Border(left=s, right=s, top=s, bottom=s)

def _orta():
    return Alignment(horizontal="center", vertical="center")

def _sol():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _baslik_satiri(ws, row, metin, cols="A1:J1", bg=C_BASLIK_BG, size=13):
    ws.merge_cells(f"A{row}:{cols.split(':')[1]}")
    c = ws[f"A{row}"]
    c.value = metin
    c.font  = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=size)
    c.fill  = _fill(bg)
    c.alignment = _orta()
    ws.row_dimensions[row].height = 30


def _satir(ws, row, degerler, bg, font=None, son_col_sol=False):
    ws.row_dimensions[row].height = 18
    for col, val in enumerate(degerler, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = font or Font(name="Calibri", size=10)
        c.fill   = _fill(bg)
        c.border = _border()
        c.alignment = _sol() if (son_col_sol and col == 1) else _orta()
        if col == 2 and isinstance(val, (int, float)) and not isinstance(val, bool):
            c.number_format = "#,##0"


def excel_export(ozet_satirlar, maliyet, prob, data, cikti_yolu):
    if not ozet_satirlar:
        print("  [UYARI] Çözüm yok, Excel oluşturulmadı.")
        return

    wb     = Workbook()
    depo_adlari = prob["depo_adlari"]
    toplam_mesafe = maliyet["toplam_mesafe"]
    toplam_yuklu  = maliyet["toplam_yuklu"]
    inhouse_count  = prob["inhouse_count"]
    outsource_count = prob["outsource_count"]

    normal_font = Font(name="Calibri", size=10)
    bold_font   = Font(name="Calibri", bold=True, size=10)
    depo_font   = Font(name="Calibri", bold=True, color=C_DEPO_FG, size=10)
    top_font    = Font(name="Calibri", bold=True, color=C_TOPLAM_FG, size=11)

    # ── SAYFA 1: Özet Filo Raporu ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Özet Filo Raporu"

    _baslik_satiri(ws1, 1, "LC WİKİKİ — MDCVRP 200 KM EŞİĞİ — ÖZET FİLO RAPORU",
                   cols="A1:J1", size=13)
    ws1.merge_cells("A2:J2")
    c2 = ws1["A2"]
    c2.value = (f"Oluşturma: {datetime.datetime.now():%d.%m.%Y %H:%M}  |  "
                f"In-house: {inhouse_count} mağaza  |  Outsource: {outsource_count} mağaza  |  "
                f"Eşik: {ESIK_KM} km")
    c2.font      = Font(name="Calibri", italic=True, color="555555", size=9)
    c2.fill      = _fill("D6E4F0")
    c2.alignment = _orta()
    ws1.row_dimensions[2].height = 15

    basliklar = ["Araç", "Çıkış Deposu", "Dönüş Deposu",
                 "Mağaza Sayısı", "Yük (koli)", "Doluluk %",
                 "Mesafe (km)", "Süre (dk)", "540dk Aşımı", "Rota (global indeksler)"]
    ws1.row_dimensions[3].height = 20
    for col, bas in enumerate(basliklar, start=1):
        c = ws1.cell(row=3, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG)
        c.alignment = _orta()
        c.border = _border()

    satir = 4
    for i, b in enumerate(ozet_satirlar):
        rota_str = " → ".join(str(g) for g in b["magaza_global"])
        bg = C_SATIR1 if i % 2 == 0 else C_SATIR2
        vals = [
            f"Araç {b['arac_no']}",
            f"Depo {b['baslangic']} – {depo_adlari[b['baslangic']]}",
            f"Depo {b['bitis']}",
            len(b["magaza_local"]),
            b["yuk"],
            b["doluluk"],
            b["mesafe_km"],
            b["sure_dk"],
            b["asim"],
            rota_str,
        ]
        ws1.row_dimensions[satir].height = 18
        for col, val in enumerate(vals, start=1):
            c = ws1.cell(row=satir, column=col, value=val)
            c.font   = normal_font
            c.fill   = _fill(bg)
            c.border = _border()
            c.alignment = _sol() if col == 10 else _orta()
        satir += 1

    # Toplam satırı
    ws1.row_dimensions[satir].height = 22
    top_vals = [f"TOPLAM ({toplam_yuklu} aktif araç)",
                "", "", sum(len(b["magaza_local"]) for b in ozet_satirlar),
                "", "", round(toplam_mesafe, 1), "", "", ""]
    for col, val in enumerate(top_vals, start=1):
        c = ws1.cell(row=satir, column=col, value=val)
        c.font = top_font; c.fill = _fill(C_TOPLAM_BG)
        c.alignment = _orta(); c.border = _border()

    genislikler = [12, 30, 14, 16, 14, 12, 16, 12, 14, 90]
    for i, g in enumerate(genislikler, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = g

    # ── SAYFA 2: Araç Detay Rotaları ──────────────────────────────────────
    ws2 = wb.create_sheet("Araç Detay Rotaları")
    _baslik_satiri(ws2, 1, "LC WİKİKİ — ARAÇ DETAY ROTA TABLOSU", cols="A1:F1", size=12)

    det_basliklar = ["Araç No", "Sıra", "Global İndeks", "Sub-Matris İndeks", "Tür", "Birikim Yük (koli)"]
    ws2.row_dimensions[2].height = 18
    for col, bas in enumerate(det_basliklar, start=1):
        c = ws2.cell(row=2, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG); c.alignment = _orta(); c.border = _border()

    det_satir = 3
    for b in ozet_satirlar:
        v_num = b["arac_no"]
        # Çıkış deposu
        for col, val in enumerate(
            [f"Araç {v_num}", 0, b["baslangic"], b["baslangic"], "DEPO (Çıkış)", 0], start=1
        ):
            c = ws2.cell(row=det_satir, column=col, value=val)
            c.font = depo_font; c.fill = _fill(C_DEPO_BG)
            c.alignment = _orta(); c.border = _border()
        ws2.row_dimensions[det_satir].height = 16
        det_satir += 1

        birikim = 0
        for sira, (local_n, global_n) in enumerate(
            zip(b["magaza_local"], b["magaza_global"]), start=1
        ):
            birikim += STORE_DEMAND
            bg = C_MAGAZA_BG if sira % 2 == 1 else C_SATIR2
            for col, val in enumerate(
                [f"Araç {v_num}", sira, global_n, local_n, "Mağaza", birikim], start=1
            ):
                c = ws2.cell(row=det_satir, column=col, value=val)
                c.font = normal_font; c.fill = _fill(bg)
                c.alignment = _orta(); c.border = _border()
            ws2.row_dimensions[det_satir].height = 15
            det_satir += 1

        # Dönüş deposu
        for col, val in enumerate(
            [f"Araç {v_num}", "-", b["bitis"], b["bitis"], "DEPO (Dönüş)", birikim], start=1
        ):
            c = ws2.cell(row=det_satir, column=col, value=val)
            c.font = depo_font; c.fill = _fill(C_DEPO_BG)
            c.alignment = _orta(); c.border = _border()
        ws2.row_dimensions[det_satir].height = 16
        det_satir += 2

    for i, g in enumerate([12, 8, 18, 18, 18, 22], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = g

    # ── SAYFA 3: Performans Özeti ──────────────────────────────────────────
    ws3 = wb.create_sheet("Performans Özeti")
    _baslik_satiri(ws3, 1, "ARAÇ PERFORMANS ÖZETİ", cols="A1:D1", size=12)

    for col, bas in enumerate(["Metrik", "Değer", "Birim", "Not"], start=1):
        c = ws3.cell(row=2, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG); c.alignment = _orta(); c.border = _border()
    ws3.row_dimensions[2].height = 18

    toplam_magaza = sum(len(b["magaza_local"]) for b in ozet_satirlar)
    ort_mesafe  = round(toplam_mesafe / len(ozet_satirlar), 1)
    ort_doluluk = round(sum(b["doluluk"]  for b in ozet_satirlar) / len(ozet_satirlar), 1)
    min_mesafe  = min(b["mesafe_km"] for b in ozet_satirlar)
    max_mesafe  = max(b["mesafe_km"] for b in ozet_satirlar)
    ort_sure    = round(sum(b["sure_dk"] for b in ozet_satirlar) / len(ozet_satirlar), 1)
    maks_sure   = max(b["sure_dk"] for b in ozet_satirlar)
    altinda     = sum(1 for b in ozet_satirlar if b["sure_dk"] <= 540)
    ustunde     = sum(1 for b in ozet_satirlar if b["sure_dk"] > 540)

    perf_veriler = [
        ("Toplam Aktif Araç",            toplam_yuklu,        "adet", ""),
        ("In-house Mağaza Sayısı",       inhouse_count,       "adet", f"Eşik: {ESIK_KM} km"),
        ("Outsource Mağaza Sayısı",      outsource_count,     "adet", ""),
        ("Ziyaret Edilen Toplam Mağaza", toplam_magaza,       "adet", "Rota bazlı"),
        ("Toplam Filo Mesafesi",         round(toplam_mesafe, 1), "km", ""),
        ("Ortalama Araç Mesafesi",       ort_mesafe,          "km",   ""),
        ("En Kısa Rota",                 min_mesafe,          "km",   ""),
        ("En Uzun Rota",                 max_mesafe,          "km",   ""),
        ("Ortalama Kapasite Doluluk",    ort_doluluk,         "%",    ""),
        ("Ortalama Rota Süresi",         ort_sure,            "dk",   ""),
        ("Maksimum Rota Süresi",         maks_sure,           "dk",   ""),
        ("540 dk Altında",               altinda,             "adet", "+15 dk/mağaza hizmet"),
        ("540 dk Üstünde",               ustunde,             "adet", ""),
        ("Zaman Limiti (OR-Tools)",      TIME_LIMIT_SECONDS,  "sn",   "GLS"),
        ("Araç Kapasitesi",              VEHICLE_CAPACITY,    "koli", ""),
        ("Mağaza Talebi",                STORE_DEMAND,        "koli", ""),
    ]

    for p_idx, (metrik, deger, birim, not_) in enumerate(perf_veriler, start=3):
        bg = C_SATIR1 if p_idx % 2 == 1 else C_SATIR2
        ws3.row_dimensions[p_idx].height = 18
        for col, val in enumerate([metrik, deger, birim, not_], start=1):
            c = ws3.cell(row=p_idx, column=col, value=val)
            c.font = normal_font; c.fill = _fill(bg)
            c.border = _border()
            c.alignment = _sol() if col == 1 else _orta()

    for i, g in enumerate([34, 16, 10, 34], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = g

    # ── SAYFA 4: Maliyet Analizi ───────────────────────────────────────────
    ws4 = wb.create_sheet("Maliyet Analizi")
    _baslik_satiri(ws4, 1, "LC WİKİKİ — MALİYET ANALİZİ (Günlük Bazda)", cols="A1:D1", size=13)

    for col, bas in enumerate(["Kalem", "Tutar (TL)", "Birim", "Not"], start=1):
        c = ws4.cell(row=2, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG); c.alignment = _orta(); c.border = _border()
    ws4.row_dimensions[2].height = 18

    def _bolum(ws, row, metin, bg=C_DEPO_BG):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=metin)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(bg); c.alignment = _sol(); c.border = _border()
        ws.row_dimensions[row].height = 18

    def _mal_satir(ws, row, kalem, tutar, birim, not_, bg=C_SATIR1, font=None):
        ws.row_dimensions[row].height = 18
        for col, val in enumerate([kalem, tutar, birim, not_], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font or normal_font
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _sol() if col == 1 else _orta()
            if col == 2 and isinstance(val, (int, float)):
                c.number_format = "#,##0"

    m = maliyet
    _bolum(ws4, 3, "  In-house Maliyet Kalemleri")
    _mal_satir(ws4, 4, "Yakıt",
               round(m["yakit"]), "TL/gün",
               f"{toplam_mesafe:.0f} km × {YAKIT_TUKETIM_LT_KM} lt × {YAKIT_FIYATI_TL} TL")
    _mal_satir(ws4, 5, "Araç Sabit",
               round(m["sabit"]), "TL/gün",
               f"{toplam_yuklu} araç × {ARAC_GUNLUK_SABIT_TL:,} TL", bg=C_SATIR2)
    _mal_satir(ws4, 6, "Sürücü",
               round(m["personel"]), "TL/gün",
               f"{toplam_yuklu} araç × {SURUCU_GUNLUK_TL:,} TL")
    _mal_satir(ws4, 7, "In-house Günlük Toplam",
               round(m["inhouse_toplam"]), "TL/gün", "",
               bg=C_TOPLAM_BG,
               font=Font(name="Calibri", bold=True, color=C_TOPLAM_FG, size=10))

    _bolum(ws4, 8, "  Outsource (3PL) Maliyet")
    _mal_satir(ws4, 9, "Pilot Tam Outsource (204 mağaza)",
               round(m["pilot_outsource"]), "TL/gün",
               f"{PILOT_MAGAZA_SAYISI} mağaza × {STORE_DEMAND} × {OUTSOURCE_KOLI_FIYAT_TL} TL")
    _mal_satir(ws4, 10, "Hibrit Outsource Kısmı",
               round(m["hibrit_out"]), "TL/gün",
               f"{outsource_count} mağaza × {STORE_DEMAND} × {OUTSOURCE_KOLI_FIYAT_TL} TL",
               bg=C_SATIR2)

    _bolum(ws4, 11, "  Karşılaştırma (Pilot 204 mağaza bazı)", bg=C_BOLUM_BG)
    _mal_satir(ws4, 12, "Mevcut Sistem — Pilot Tam 3PL",
               round(m["pilot_outsource"]), "TL/gün", f"{PILOT_MAGAZA_SAYISI} mağaza")
    _mal_satir(ws4, 13, "Önerilen Hibrit Sistem",
               round(m["hibrit_toplam"]), "TL/gün",
               f"{inhouse_count} in-house + {outsource_count} outsource", bg=C_SATIR2)
    _mal_satir(ws4, 14, "Günlük Tasarruf",
               round(m["tasarruf_gun"]), "TL/gün", "Pilot Tam 3PL – Hibrit",
               bg=C_KAZANC_BG,
               font=Font(name="Calibri", bold=True, color=C_KAZANC_FG, size=10))
    _mal_satir(ws4, 15, "Yıllık Tasarruf (300 iş günü)",
               round(m["tasarruf_yil"]), "TL/yıl", "Günlük × 300",
               bg=C_KAZANC_BG,
               font=Font(name="Calibri", bold=True, color=C_KAZANC_FG, size=11))
    _mal_satir(ws4, 16, "Tasarruf Oranı",
               f"%{m['tasarruf_oran']:.1f}", "", "Pilot 204 mağaza bazı",
               bg=C_KAZANC_BG,
               font=Font(name="Calibri", bold=True, color=C_KAZANC_FG, size=10))
    _bolum(ws4, 17, "  Ham Referans — Kapsam Uyumsuz (Kullanılmaz)", bg="808080")
    _mal_satir(ws4, 18, "Tüm LCW Sistemi (533 mağaza)",
               round(m["tam_outsource_ham"]), "TL/gün",
               f"{TOPLAM_MAGAZA_HAM} mağaza × {STORE_DEMAND} × {OUTSOURCE_KOLI_FIYAT_TL} TL")

    for i, g in enumerate([38, 20, 12, 50], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = g

    # ── SAYFA 5: Duyarlılık Analizi ───────────────────────────────────────
    ws5 = wb.create_sheet("Duyarlılık Analizi")
    _baslik_satiri(ws5, 1, "LC WİKİKİ — DUYARLILIK ANALİZİ", cols="A1:F1", size=13)

    ws5.merge_cells("A2:F2")
    c2 = ws5["A2"]
    c2.value = (f"Taban: 3PL={OUTSOURCE_KOLI_FIYAT_TL} TL/koli  |  "
                f"Yakıt={YAKIT_FIYATI_TL} TL/lt  |  "
                f"In-house mesafe={maliyet['toplam_mesafe']:.0f} km  |  "
                f"Aktif araç={maliyet['toplam_yuklu']}")
    c2.font = Font(name="Calibri", italic=True, color="555555", size=9)
    c2.fill = _fill("D6E4F0"); c2.alignment = _orta()
    ws5.row_dimensions[2].height = 15

    C_TABAN   = "FFE699"   # altın sarı — taban değer satırı
    C_NEGATIF = "FFCCCC"   # açık kırmızı — negatif tasarruf

    satir = 3

    # ── Senaryo 1: 3PL Koli Fiyatı ────────────────────────────────────────
    _bolum(ws5, satir, "  SENARYO 1 — 3PL Koli Fiyatı Duyarlılığı (In-house maliyet sabit)", bg=C_BOLUM_BG)
    satir += 1

    s1_bas = ["3PL Fiyatı (TL/koli)", "Tam Outsource (TL/gün)",
              "Hibrit Out Kısmı (TL/gün)", "Hibrit Toplam (TL/gün)",
              "Günlük Tasarruf (TL)", "Tasarruf Oranı %"]
    ws5.row_dimensions[satir].height = 18
    for col, bas in enumerate(s1_bas, start=1):
        c = ws5.cell(row=satir, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG); c.alignment = _orta(); c.border = _border()
    satir += 1

    s1_fiyatlar = [3.00, 3.38, 3.75, 4.13, 4.50]
    inhouse_sabit = maliyet["inhouse_toplam"]
    for fiyat in s1_fiyatlar:
        tam_out_f   = PILOT_MAGAZA_SAYISI * STORE_DEMAND * fiyat
        hibrit_of   = outsource_count * STORE_DEMAND * fiyat
        hibrit_f    = inhouse_sabit + hibrit_of
        tasarruf_f  = tam_out_f - hibrit_f
        oran_f      = tasarruf_f / tam_out_f * 100
        is_taban    = abs(fiyat - OUTSOURCE_KOLI_FIYAT_TL) < 0.01
        bg = C_TABAN if is_taban else (C_NEGATIF if tasarruf_f < 0 else (C_SATIR1 if satir % 2 == 1 else C_SATIR2))
        font = Font(name="Calibri", bold=True, size=10) if is_taban else Font(name="Calibri", size=10)
        ws5.row_dimensions[satir].height = 18
        for col, val in enumerate([fiyat, round(tam_out_f), round(hibrit_of),
                                    round(hibrit_f), round(tasarruf_f), f"%{oran_f:.1f}"], start=1):
            c = ws5.cell(row=satir, column=col, value=val)
            c.font = font; c.fill = _fill(bg); c.alignment = _orta(); c.border = _border()
            if col in (1,) and isinstance(val, float):
                c.number_format = "0.00"
            elif col in (2, 3, 4, 5) and isinstance(val, int):
                c.number_format = "#,##0"
        satir += 1

    ws5.merge_cells(f"A{satir}:F{satir}")
    cnot = ws5[f"A{satir}"]
    cnot.value = ("Not: Pilot 204 mağaza bazında tasarruf oranı bandı: "
                  "%4.6 (3.00 TL) – %19.2 (4.50 TL). Taban 3.75 TL'de %13.4.")
    cnot.font = Font(name="Calibri", italic=True, color="444444", size=9)
    cnot.fill = _fill("F2F2F2")
    cnot.alignment = _sol()
    ws5.row_dimensions[satir].height = 15
    satir += 1

    satir += 1  # boşluk

    # ── Senaryo 2: Yakıt Fiyatı ───────────────────────────────────────────
    _bolum(ws5, satir, "  SENARYO 2 — Yakıt Fiyatı Duyarlılığı (Outsource maliyet sabit)", bg=C_BOLUM_BG)
    satir += 1

    s2_bas = ["Yakıt Fiyatı (TL/lt)", "In-house Yakıt (TL/gün)",
              "In-house Toplam (TL/gün)", "Hibrit Toplam (TL/gün)",
              "Günlük Tasarruf (TL)", "Tasarruf Oranı %"]
    ws5.row_dimensions[satir].height = 18
    for col, bas in enumerate(s2_bas, start=1):
        c = ws5.cell(row=satir, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG); c.alignment = _orta(); c.border = _border()
    satir += 1

    s2_fiyatlar = [34.4, 38.7, 43.0, 47.3, 51.6]
    hibrit_out_sabit  = maliyet["hibrit_out"]
    tam_out_sabit     = maliyet["pilot_outsource"]
    for yakit_f in s2_fiyatlar:
        yakit_maliyet = maliyet["toplam_mesafe"] * YAKIT_TUKETIM_LT_KM * yakit_f
        inhouse_f     = yakit_maliyet + maliyet["sabit"] + maliyet["personel"]
        hibrit_f      = inhouse_f + hibrit_out_sabit
        tasarruf_f    = tam_out_sabit - hibrit_f
        oran_f        = tasarruf_f / tam_out_sabit * 100
        is_taban      = abs(yakit_f - YAKIT_FIYATI_TL) < 0.01
        bg = C_TABAN if is_taban else (C_NEGATIF if tasarruf_f < 0 else (C_SATIR1 if satir % 2 == 1 else C_SATIR2))
        font = Font(name="Calibri", bold=True, size=10) if is_taban else Font(name="Calibri", size=10)
        ws5.row_dimensions[satir].height = 18
        for col, val in enumerate([yakit_f, round(yakit_maliyet), round(inhouse_f),
                                    round(hibrit_f), round(tasarruf_f), f"%{oran_f:.1f}"], start=1):
            c = ws5.cell(row=satir, column=col, value=val)
            c.font = font; c.fill = _fill(bg); c.alignment = _orta(); c.border = _border()
            if col == 1 and isinstance(val, float):
                c.number_format = "0.0"
            elif col in (2, 3, 4, 5) and isinstance(val, int):
                c.number_format = "#,##0"
        satir += 1

    satir += 1  # boşluk

    # ── Senaryo 3: Geri Ödeme Süresi ──────────────────────────────────────
    _bolum(ws5, satir, "  SENARYO 3 — Geri Ödeme Süresi (Payback Period)", bg=C_BOLUM_BG)
    satir += 1

    ARAC_YATIRIM    = 5_000_000
    toplam_yatirim  = maliyet["toplam_yuklu"] * ARAC_YATIRIM
    yillik_tasarruf = maliyet["tasarruf_gun"] * 300
    payback_yil     = toplam_yatirim / yillik_tasarruf if yillik_tasarruf > 0 else float("inf")
    payback_ay      = payback_yil * 12
    payback_durum   = "✓ 3 YIL ALTINDA" if payback_yil < 3 else "✗ 3 YIL ÜSTÜNDE"

    s3_veriler = [
        ("Araç Başı Yatırım Maliyeti",  ARAC_YATIRIM,          "TL/araç",  "Yeni kamyon varsayımı"),
        ("Aktif Araç Sayısı",            maliyet["toplam_yuklu"],"adet",     ""),
        ("Toplam Araç Yatırımı",         toplam_yatirim,        "TL",       f"{maliyet['toplam_yuklu']} araç × {ARAC_YATIRIM:,} TL"),
        ("Yıllık Net Tasarruf",          round(yillik_tasarruf),"TL/yıl",   "Günlük × 300 iş günü"),
        ("Geri Ödeme Süresi",            round(payback_yil, 2), "yıl",      f"≈ {payback_ay:.1f} ay"),
        ("3 Yıl Eşiği Değerlendirme",    payback_durum,         "",         "3 yıl = endüstri standardı"),
    ]

    ws5.row_dimensions[satir].height = 18
    for col, bas in enumerate(["Kalem", "Değer", "Birim", "Not"], start=1):
        c = ws5.cell(row=satir, column=col, value=bas)
        c.font = Font(name="Calibri", bold=True, color=C_BASLIK_FG, size=10)
        c.fill = _fill(C_DEPO_BG); c.alignment = _orta(); c.border = _border()
    satir += 1

    for idx, (kalem, deger, birim, not_) in enumerate(s3_veriler):
        is_payback = "Geri Ödeme" in kalem
        is_durum   = "Eşiği" in kalem
        if is_durum:
            bg   = C_KAZANC_BG if payback_yil < 3 else C_NEGATIF
            font = Font(name="Calibri", bold=True,
                        color=C_KAZANC_FG if payback_yil < 3 else "CC0000", size=10)
        elif is_payback:
            bg   = C_TABAN
            font = Font(name="Calibri", bold=True, size=10)
        else:
            bg   = C_SATIR1 if idx % 2 == 0 else C_SATIR2
            font = Font(name="Calibri", size=10)

        ws5.row_dimensions[satir].height = 20
        for col, val in enumerate([kalem, deger, birim, not_], start=1):
            c = ws5.cell(row=satir, column=col, value=val)
            c.font = font; c.fill = _fill(bg); c.border = _border()
            c.alignment = _sol() if col == 1 else _orta()
            if col == 2 and isinstance(val, (int, float)) and not isinstance(val, bool):
                c.number_format = "#,##0" if isinstance(val, int) else "0.00"
        satir += 1

    for i, g in enumerate([38, 22, 14, 42], start=1):
        ws5.column_dimensions[get_column_letter(i)].width = g

    wb.save(cikti_yolu)
    print(f"\n  Excel raporu kaydedildi: {cikti_yolu}")


# ─────────────────────────────────────────────────────────────────────────────
# 6. ANA PROGRAM
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print()
    print("=" * 72)
    print("  LC WİKİKİ — MDCVRP 200 KM EŞİĞİ (OR-Tools SAVINGS + GLS)")
    print("=" * 72)

    print("\n[1/4] Veri hazırlanıyor...")
    prob = hazirla_alt_problem()

    n_sub = prob["n_sub"]
    print(f"\n  Alt matris boyutu : {n_sub} × {n_sub}")
    print(f"  (6 depo + {n_sub - NUM_DEPOTS} mağaza)")

    print("\n[2/4] OR-Tools modeli oluşturuluyor...")
    data = veri_modeli_olustur(prob)
    print(f"  Toplam düğüm    : {data['num_locations']}")
    print(f"  Araç sayısı     : {data['num_vehicles']}")
    print(f"  Starts/Ends     : {data['starts']} / {data['ends']}")

    print("\n[3/4] Optimizasyon çalıştırılıyor...")
    manager, routing, solution = coz(data)

    print("\n[4/4] Raporlar hazırlanıyor...")
    sonuc = rapor_yazdir(manager, routing, solution, data, prob)
    if sonuc is None:
        return

    ozet_satirlar, maliyet = sonuc
    excel_yolu = OUTPUT_DIR / "SONUC_rota_maliyet.xlsx"
    excel_export(ozet_satirlar, maliyet, prob, data, excel_yolu)


if __name__ == "__main__":
    main()
