#!/usr/bin/env python3
"""
LCW-Matrix v1.1 — Depolu Versiyon
Girdi : LCW_Lokasyonlar.xlsx  (6 depo + 204 mağaza = 210 satır)
Çıktı : LCW_Mesafe_Sure_Matrisi.xlsx  (210×210 matris)
Motor : OSRM Table API (kamu sunucusu, sürüş profili)
"""
from __future__ import annotations

import time
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ─── YAPILANDIRMA ─────────────────────────────────────────────────────────────

BASE_DIR    = Path(__file__).resolve().parent.parent / "01_Veri"
INPUT_FILE  = BASE_DIR / "LCW_Lokasyonlar.xlsx"
OUTPUT_FILE = BASE_DIR / "LCW_Mesafe_Sure_Matrisi.xlsx"
LOG_DIR     = Path(__file__).resolve().parent.parent / "logs"

OSRM_BASE_URL   = "http://router.project-osrm.org/table/v1/driving"
BATCH_SIZE      = 10
REQUEST_TIMEOUT = 120
MAX_RETRIES     = 3
RETRY_WAIT      = 15

# ─── LOGGING ──────────────────────────────────────────────────────────────────

def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    fh  = logging.FileHandler(LOG_DIR / f"matrix_depolu_{ts}.log", encoding="utf-8")
    sh  = logging.StreamHandler()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    fh.setFormatter(fmt); sh.setFormatter(fmt)
    logger = logging.getLogger("lcw-matrix-depolu")
    logger.setLevel(logging.INFO)
    logger.addHandler(fh); logger.addHandler(sh)
    return logger


# ─── VERİ OKUMA ───────────────────────────────────────────────────────────────

def load_locations(path: Path, logger: logging.Logger) -> pd.DataFrame:
    logger.info(f"Excel okunuyor: {path}")
    df = pd.read_excel(path, sheet_name=0)

    required = {"Mağaza Adı", "İl", "İlçe", "Enlem", "Boylam"}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(f"Excel'de eksik sütun(lar): {missing}")

    before = len(df)
    df = df.dropna(subset=["Enlem", "Boylam"]).reset_index(drop=True)
    if len(df) < before:
        logger.warning(f"{before - len(df)} satır koordinat eksikliği nedeniyle atlandı.")

    logger.info(f"{len(df)} lokasyon yüklendi.")
    return df


# ─── ETİKET ÜRETİCİ ───────────────────────────────────────────────────────────

def build_labels(df: pd.DataFrame) -> list[str]:
    labels = []
    for i, row in df.iterrows():
        ilce = str(row["İlçe"]).strip() if pd.notna(row["İlçe"]) else ""
        if ilce and ilce.lower() != "nan":
            labels.append(f"{i+1}. {row['İl']} — {ilce}")
        else:
            labels.append(f"{i+1}. {row['İl']}")
    return labels


# ─── OSRM İSTEMCİSİ ───────────────────────────────────────────────────────────

class OsrmClient:
    def __init__(self, base_url: str, logger: logging.Logger):
        self.base_url = base_url.rstrip("/")
        self.logger   = logger
        self.total_requests = 0

    @staticmethod
    def _coords_str(df: pd.DataFrame) -> str:
        parts = [f"{row['Boylam']},{row['Enlem']}" for _, row in df.iterrows()]
        return ";".join(parts)

    def _get(self, url: str, params: dict, attempt: int = 1) -> Optional[dict]:
        try:
            resp = requests.get(url, params=params, timeout=REQUEST_TIMEOUT)
            self.total_requests += 1

            if resp.status_code == 200:
                data = resp.json()
                if data.get("code") == "Ok":
                    return data
                self.logger.error(f"OSRM hata kodu: {data.get('code')} — {data.get('message')}")
                return None

            if resp.status_code == 429:
                if attempt <= MAX_RETRIES:
                    self.logger.warning(f"Rate limit. {RETRY_WAIT}s bekleniyor... (Deneme {attempt})")
                    time.sleep(RETRY_WAIT)
                    return self._get(url, params, attempt + 1)
                self.logger.error("Rate limit: Maksimum deneme aşıldı.")
                return None

            if resp.status_code >= 500:
                if attempt <= MAX_RETRIES:
                    self.logger.warning(f"Sunucu hatası ({resp.status_code}). {RETRY_WAIT}s bekleniyor...")
                    time.sleep(RETRY_WAIT)
                    return self._get(url, params, attempt + 1)
                self.logger.error(f"Sunucu hatası: {resp.status_code}")
                return None

            self.logger.error(f"Beklenmeyen HTTP {resp.status_code}")
            return None

        except requests.exceptions.Timeout:
            if attempt <= MAX_RETRIES:
                self.logger.warning(f"Zaman aşımı ({REQUEST_TIMEOUT}s). Deneme {attempt}/{MAX_RETRIES}...")
                time.sleep(RETRY_WAIT)
                return self._get(url, params, attempt + 1)
            self.logger.error("Zaman aşımı: Maksimum deneme aşıldı.")
            return None

        except requests.exceptions.ConnectionError:
            if attempt <= MAX_RETRIES:
                self.logger.warning(f"Bağlantı hatası. Deneme {attempt}/{MAX_RETRIES}...")
                time.sleep(RETRY_WAIT)
                return self._get(url, params, attempt + 1)
            self.logger.error("Bağlantı hatası: Maksimum deneme aşıldı.")
            return None

    def compute_matrix(
        self, df: pd.DataFrame
    ) -> tuple[list[list[Optional[float]]], list[list[Optional[float]]]]:
        n          = len(df)
        coords_str = self._coords_str(df)
        all_dest   = ";".join(str(i) for i in range(n))
        url        = f"{self.base_url}/{coords_str}"

        dist_matrix: list[list[Optional[float]]] = [[None] * n for _ in range(n)]
        dur_matrix:  list[list[Optional[float]]] = [[None] * n for _ in range(n)]

        source_batches = [
            list(range(start, min(start + BATCH_SIZE, n)))
            for start in range(0, n, BATCH_SIZE)
        ]

        total_batches = len(source_batches)
        self.logger.info(
            f"Matris hesabı: {n}×{n} = {n*n:,} hücre | "
            f"{total_batches} grup × {BATCH_SIZE} kaynak | "
            f"Sunucu: {self.base_url}"
        )

        for batch_idx, source_indices in enumerate(source_batches, 1):
            src_str = ";".join(str(i) for i in source_indices)
            params  = {
                "sources":      src_str,
                "destinations": all_dest,
                "annotations":  "duration,distance",
            }

            self.logger.info(
                f"  Grup {batch_idx}/{total_batches}: "
                f"kaynak {source_indices[0]}–{source_indices[-1]} "
                f"→ {len(source_indices)} × {n} çift hesaplanıyor..."
            )

            data = self._get(url, params)

            if data is None:
                self.logger.error(
                    f"  Grup {batch_idx} başarısız — bu kaynaklar için değerler None kalacak."
                )
                continue

            durations = data.get("durations", [])
            distances = data.get("distances", [])

            for local_idx, global_src in enumerate(source_indices):
                if local_idx < len(durations):
                    for dst, val in enumerate(durations[local_idx]):
                        if val is not None:
                            dur_matrix[global_src][dst] = round(val / 60, 1)

                if local_idx < len(distances):
                    for dst, val in enumerate(distances[local_idx]):
                        if val is not None:
                            dist_matrix[global_src][dst] = round(val / 1000, 2)

            self.logger.info(f"  Grup {batch_idx}/{total_batches} tamamlandı.")

            if batch_idx < total_batches:
                time.sleep(2)

        return dist_matrix, dur_matrix


# ─── EXCEL DIŞA AKTARICISI ────────────────────────────────────────────────────

class MatrixExporter:
    HDR_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HDR_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
    LABEL_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    DIAG_FILL  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    THIN       = Side(style="thin", color="BFBFBF")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def _write_matrix_sheet(
        self,
        wb:     Workbook,
        title:  str,
        unit:   str,
        matrix: list[list[Optional[float]]],
        labels: list[str],
        df:     pd.DataFrame,
    ) -> None:
        ws = wb.create_sheet(title)
        n  = len(labels)

        corner = ws.cell(row=1, column=1, value=f"{title}\n({unit})")
        corner.fill      = self.HDR_FILL
        corner.font      = self.HDR_FONT
        corner.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        for j, lbl in enumerate(labels):
            c = ws.cell(row=1, column=j + 2, value=lbl)
            c.fill      = self.HDR_FILL
            c.font      = self.HDR_FONT
            c.alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")

        for i, lbl in enumerate(labels):
            rc = ws.cell(row=i + 2, column=1, value=lbl)
            rc.fill      = self.LABEL_FILL
            rc.font      = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
            rc.alignment = Alignment(horizontal="left", vertical="center")

            for j in range(n):
                val = matrix[i][j]
                c   = ws.cell(row=i + 2, column=j + 2, value=val)
                c.border    = self.BORDER
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font      = Font(name="Calibri", size=9)
                if i == j:
                    c.fill  = self.DIAG_FILL
                    c.value = 0

        ws.column_dimensions["A"].width = 30
        for j in range(n):
            ws.column_dimensions[get_column_letter(j + 2)].width = 8

        ws.row_dimensions[1].height = 80
        for i in range(n):
            ws.row_dimensions[i + 2].height = 14

        ws.freeze_panes = "B2"

        data_range = f"B2:{get_column_letter(n + 1)}{n + 1}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="num",      start_value=0,   start_color="63BE7B",
                mid_type="percentile", mid_value=50,    mid_color="FFEB84",
                end_type="max",                          end_color="F8696B",
            ),
        )

    def _write_summary_sheet(
        self,
        wb:          Workbook,
        df:          pd.DataFrame,
        dist_matrix: list[list[Optional[float]]],
        dur_matrix:  list[list[Optional[float]]],
        api_calls:   int,
    ) -> None:
        ws = wb.create_sheet("Özet")
        n  = len(df)

        flat_dist = [
            dist_matrix[i][j]
            for i in range(n) for j in range(n)
            if i != j and dist_matrix[i][j] is not None
        ]
        flat_dur = [
            dur_matrix[i][j]
            for i in range(n) for j in range(n)
            if i != j and dur_matrix[i][j] is not None
        ]

        none_count = sum(
            1 for i in range(n) for j in range(n)
            if i != j and dist_matrix[i][j] is None
        )

        ws.append(["Metrik", "Değer"])
        for c in ws[1]:
            c.fill = self.HDR_FILL; c.font = self.HDR_FONT

        rows = [
            ["Lokasyon Sayısı",              n],
            ["Toplam Çift Sayısı",           n * (n - 1)],
            ["Hesaplanamayan Çift",           none_count],
            ["Ortalama Mesafe (km)",          round(sum(flat_dist) / len(flat_dist), 2) if flat_dist else "—"],
            ["Minimum Mesafe (km)",           round(min(flat_dist), 2) if flat_dist else "—"],
            ["Maksimum Mesafe (km)",          round(max(flat_dist), 2) if flat_dist else "—"],
            ["Ortalama Süre (dk)",            round(sum(flat_dur)  / len(flat_dur),  1) if flat_dur  else "—"],
            ["Minimum Süre (dk)",             round(min(flat_dur),  1) if flat_dur  else "—"],
            ["Maksimum Süre (dk)",            round(max(flat_dur),  1) if flat_dur  else "—"],
            ["Toplam OSRM İsteği",            api_calls],
            ["Hesaplama Tarihi",              datetime.now().isoformat(timespec="seconds")],
            ["OSRM Sunucusu",                 OSRM_BASE_URL],
        ]
        for row in rows:
            ws.append(row)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 35

    def export(
        self,
        df:          pd.DataFrame,
        labels:      list[str],
        dist_matrix: list[list[Optional[float]]],
        dur_matrix:  list[list[Optional[float]]],
        api_calls:   int,
        output_path: Path,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        self._write_matrix_sheet(wb, "Mesafe (km)", "km", dist_matrix, labels, df)
        self._write_matrix_sheet(wb, "Süre (dk)",   "dk", dur_matrix,  labels, df)
        self._write_summary_sheet(wb, df, dist_matrix, dur_matrix, api_calls)

        wb.save(output_path)
        self.logger.info(f"Excel kaydedildi: {output_path}")


# ─── ANA ORKESTRATÖR ──────────────────────────────────────────────────────────

def main() -> None:
    logger = setup_logging()
    logger.info("=" * 62)
    logger.info("LCW-Matrix v1.1 (Depolu) başlatıldı")
    logger.info("=" * 62)

    logger.info("\n[ADIM 1/4] Lokasyonlar okunuyor...")
    df     = load_locations(INPUT_FILE, logger)
    labels = build_labels(df)
    n      = len(df)
    logger.info(f"[ADIM 1 TAMAM] {n} lokasyon, {n*n:,} matris hücresi hedeflendi.")

    logger.info("\n[ADIM 2/4] OSRM matris hesabı başlıyor...")
    t0     = time.time()
    client = OsrmClient(OSRM_BASE_URL, logger)
    dist_matrix, dur_matrix = client.compute_matrix(df)
    elapsed = time.time() - t0
    logger.info(f"[ADIM 2 TAMAM] Matris hesabı {elapsed:.1f}s sürdü. "
                f"Toplam OSRM isteği: {client.total_requests}")

    logger.info("\n[ADIM 3/4] Kalite kontrolü...")
    none_count  = sum(
        1 for i in range(n) for j in range(n)
        if i != j and dist_matrix[i][j] is None
    )
    total_pairs = n * (n - 1)
    coverage    = 100 * (1 - none_count / total_pairs) if total_pairs else 100
    logger.info(
        f"Kapsam: {total_pairs - none_count:,}/{total_pairs:,} çift "
        f"({coverage:.1f}%) — {none_count} hesaplanamayan çift."
    )

    logger.info(f"\n[ADIM 4/4] Excel dışa aktarım → {OUTPUT_FILE}")
    MatrixExporter(logger).export(
        df          = df,
        labels      = labels,
        dist_matrix = dist_matrix,
        dur_matrix  = dur_matrix,
        api_calls   = client.total_requests,
        output_path = OUTPUT_FILE,
    )

    flat_dist = [
        dist_matrix[i][j]
        for i in range(n) for j in range(n)
        if i != j and dist_matrix[i][j] is not None
    ]
    flat_dur = [
        dur_matrix[i][j]
        for i in range(n) for j in range(n)
        if i != j and dur_matrix[i][j] is not None
    ]

    logger.info("\n" + "=" * 62)
    logger.info("İŞLEM BAŞARIYLA TAMAMLANDI")
    logger.info(f"  Lokasyon sayısı          : {n}")
    logger.info(f"  Toplam çift              : {total_pairs:,}")
    logger.info(f"  Hesaplanan çift          : {total_pairs - none_count:,}  ({coverage:.1f}%)")
    if flat_dist:
        logger.info(f"  Mesafe ort/min/maks (km) : "
                    f"{sum(flat_dist)/len(flat_dist):.1f} / "
                    f"{min(flat_dist):.1f} / {max(flat_dist):.1f}")
    if flat_dur:
        logger.info(f"  Süre ort/min/maks (dk)   : "
                    f"{sum(flat_dur)/len(flat_dur):.1f} / "
                    f"{min(flat_dur):.1f} / {max(flat_dur):.1f}")
    logger.info(f"  OSRM istek sayısı        : {client.total_requests}")
    logger.info(f"  Toplam süre              : {elapsed:.1f}s")
    logger.info(f"  Çıktı dosyası            : {OUTPUT_FILE}")
    logger.info("=" * 62)


if __name__ == "__main__":
    main()
